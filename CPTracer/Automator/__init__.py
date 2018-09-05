from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver import Firefox
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

from os import devnull
from os import system
# from win32com.client.gencache import EnsureDispatch #@UnresolvedImport
# from win32com.client import Dispatch #@UnresolvedImport
from openpyxl import load_workbook
from openpyxl.styles.colors import RGB
from openpyxl.styles import PatternFill
from copy import copy

from sys import path
from sys import exc_info
from time import sleep

from psutil import process_iter #@UnresolvedImport

from selenium.webdriver.common.action_chains import ActionChains

from tkinter import Button, Tk, Label, Entry

import re
from pyHook.HookManager import GetKeyState
from selenium.webdriver.support.wait import WebDriverWait

# from enum.IntFlag import  



class constants:
    CONTAINERCOL = ''
    LOADCOL = ""
    SIZECOL=""
    LECOL=''
    SHIPPOINTCOL=""
    CURRENTPOSCOL=""
    LASTREPORTEDCOL=""
    STATUSCOL=""
    CONSPOINTCOL=""
    GROUNDECOL=""
    ACTARRIVALCOL=""
    HOLDSCOL=""
    LFDCOL=""
    STORAGECOL=""
    PIECESCOL=""
    PUNUMBERCOL=""


class Container:
    def __init__(self, number = ""):
        self.number = number
        self.load = ""
        self.size = ""
        self.le = ""
        self.shippoint = ""
        self.currentpos = ""
        self.lastreported = ""
        self.status =""
        self.conspoint =""
        self.grounded= ""
        self.actarrival=""
        self.holds=""
        self.lfd=""
        self.storage=""
        self.pieces=""
        self.punumber=""


def setupCP():
    fp = FirefoxProfile();
    fp.set_preference("webdriver.load.strategy", "unstable");
#     fp.set_preference("XRE_NO_WINDOWS_CRASH_DIALOG=1")
     
    driver = Firefox(firefox_profile=fp, log_path=devnull)
    driver.get("https://www8.cpr.ca/cpcustomerstation/")
#     driver.set_window_position(1920, 0)
#     sleep(30)
    driver.maximize_window()
    
    driver.implicitly_wait(40)
    
    f=open(r"J:\LOCAL DEPARTMENT\Automation - DO NOT MOVE\CP Login.txt", 'r')
    read = f.readline()
    m = re.search("username: *", read)
    username = read[m.end():].rstrip()
    read = f.readline()
    m = re.search("password: *", read)
    password = read[m.end():].rstrip()
    f.close()    
    
#     driver.find_element_by_class_name("lbl").click()
    
    driver.find_element_by_id("username").send_keys(username)
    driver.find_element_by_id("password").send_keys(password)
    driver.find_element_by_class_name("login_button").click()
    
#     driver.get("https://www.cprintermodal.ca/customer/Home.do")
#     if not "Canadian Pacific Intermodal Tools" in driver.page_source:
#         WebDriverWait(driver, 100000).until(lambda driver: "Remote Reporting Facility Login" in driver.page_source)
    driver.get("https://www.cprintermodal.ca/customer/Home.do")
#     driver.get("https://www.cprintermodal.ca/customer/Home.do")
#     https://www.cprintermodal.ca/customer/LoadTracing.do;jsessionid=CB390C00AEA401051B18CDE17D090E06.cpr-www1-prod
#     https://www.cprintermodal.ca/customer/LoadTracing.do;jsessionid=3CE125EC6173E92A9B4B25F9E7054366.cpr-www1-prod
#     driver.switch_to_frame(driver.find_element_by_css_selector("frame[src='/cpcustomerstation/mainPage']"))
#     driver.find_element_by_id("dijit_PopupMenuBarItem_1_text").click()
#     driver.find_element_by_id("dijit_MenuItem_20_text").click()
#     sleep(1)
    driver.find_element_by_css_selector("a[href*='/customer/LoadTracing.do;']").click()
    return driver
        
def readContainerInfo(driver, containers):
#     driver.switch_to.frame("content3")
#     driver.switch_to.frame("main")
#     driver.switch_to_frame(driver.find_element_by_css_selector(""))
    
    containerText = driver.find_element_by_name("paramValue3470")
    containerText.click()
    for container in containers:
        containerText.send_keys(container.number + " ")
    
    driver.find_element_by_name("paramValue3478").click()
    driver.find_element_by_name("paramValue3479").click()
    driver.find_element_by_name("paramValue3518").click()
    driver.find_element_by_name("paramValue3524").click()
    driver.find_element_by_css_selector("input[value='Run']").click()
    
    
#     driver.find_element_by_css_selector("tbody")
    
#     elem = driver.find_element_by_css_selector("img[src='/customer/images/button-next.gif'")
    lastPage = "false"
    k=0
    while lastPage=="false":
#         sleep(200)
#         while not GetKeyState(45)<0:
#             True
        i=0
        elem = driver.find_element_by_css_selector("img[src='/customer/images/button-next.gif'")
        lastPage = elem.get_attribute("aria-disabled")
        rows = driver.find_element_by_css_selector("table[class='tablesorter1 tablesorter tablesorter-default']>tbody").find_elements_by_css_selector("tr")
        for row in rows:
            if i<k*25:
                i+=1
                continue
            cells = row.find_elements_by_css_selector("td")
            container = containers[i]
#             print(i)
            i+=1
#             for Xcontainer in containers:
#                 if Xcontainer.number[:4] + " " + "000" + Xcontainer.number[4:] in cells[1].text: 
#                     container=Xcontainer
#                     break
            j = 0
            if container != "":
                for cell in cells:
                    if j==0:
                        container.load = cell.text
                    elif j==1:
                        container.size = cell.text
#                         print(cell.text)
                    elif j==2:
                        container.le = cell.text
                    elif j==3:
                        container.shippoint = cell.text
                    elif j==4:
                        container.currentpos = cell.text
                    elif j==5:
                        container.lastreported = cell.text
                    elif j==6:
                        container.status = cell.text
                    elif j==7:
                        container.conspoint = cell.text
                    elif j==8:
                        container.grounded = cell.text
                    elif j==9:
                        container.actarrival = cell.text
                    elif j==10:
                        container.holds = cell.text
                    elif j==11:
                        container.lfd = cell.text
                    elif j==12:
                        container.storage = cell.text
                    elif j==13:
                        container.pieces = cell.text
                    elif j==14:
                        container.punumber = cell.text
                    j+=1
            if i % 25==0 and not i==k*25:
                k = k+1
                break
            
        if lastPage=="false":
            elem.click()
        
#                 else:
#                     print(container)
#                 print(cell.get_attribute("value"))
#                 print(cell.tag_name)
        
#     sleep(600)    
    driver.quit()        
        
def putInfoinExcel(containers, localContainers):
    i = 0
    for row in localContainers.rows:
        if i != 0:
            if row[constants.CONTAINERCOL].value != None and row[constants.CONTAINERCOL].value != "":
#                 container = containers[i-1]
                for Xcontainer in containers:
                    if Xcontainer.number==row[constants.CONTAINERCOL].value.strip():
                        container=Xcontainer
                        break
                    
                backupRow = []
                j=0
                for _ in row:
                    if row[j].value == "":
                        backupRow.append(None)
                    else:
                        backupRow.append(row[j].value)
                    j+=1
                
                size=""
                
                if "x20'" in container.size:    
                    size="20"
                elif "x40'" in container.size:    
                    size="40"
                elif "x45'" in container.size:    
                    size="45"
                elif "x53'" in container.size:    
                    size="53"
                
                if "Dry" in container.size:
                    size=size+"D"
                elif "Open Top" in container.size:
                    size=size+"O"
                elif "Reef" in container.size:
                    size=size+"R"
                
                if "8'6\"" in container.size:
                    size=size+"86" 
                elif "9'6\"" in container.size:
                    size=size+"96"
                row[constants.SIZECOL].value=size
                    
                row[constants.LOADCOL].value=container.load
                
                    
                    
                    
                row[constants.LECOL].value=container.le
                row[constants.SHIPPOINTCOL].value=container.shippoint
                row[constants.CURRENTPOSCOL].value=container.currentpos
                row[constants.LASTREPORTEDCOL].value=container.lastreported
                row[constants.STATUSCOL].value=container.status
                row[constants.CONSPOINTCOL].value=container.conspoint
                row[constants.GROUNDECOL].value=container.grounded
                row[constants.ACTARRIVALCOL].value=container.actarrival
                row[constants.HOLDSCOL].value=container.holds
                row[constants.LFDCOL].value=container.lfd
                row[constants.STORAGECOL].value=container.storage
                row[constants.PUNUMBERCOL].value=container.punumber
                row[constants.PIECESCOL].value=container.pieces
                
#                 malPort = False
#                 hold = False
                for i in range(len(row)):
                    
        #             if backupRow[i].value != None and backupRow[i].value != row[i].value:
#                     if i>len(row)-4:
#                         print(backupRow[i])
#                         print(row[i].value)
#                         print(backupRow[i]==row[i].value)
                    if i>9:
                        if backupRow[i] != row[i].value and not (row[i].value=="" and backupRow[i]==None):
                            row[i].fill = PatternFill(start_color='CCFFFF',
                            end_color='CCFFFF',
                            fill_type='solid')
                        else:
                            row[i].fill = PatternFill(start_color='FFFFFF',
                            end_color='FFFFFF',
                            fill_type='solid')
                        
#                         if row[i].value and "Malport" in row[i].value:
#                             malPort = True
#                 if row[constants.CUSTOMSCOL].value and "Hold" in row[constants.CUSTOMSCOL].value and row[constants.ETACOL].value and row[constants.ETACOL].value != " ":
#                     hold =True
                hold = False
                if row[constants.HOLDSCOL].value and not "none" in row[constants.HOLDSCOL].value.lower() and (row[constants.ACTARRIVALCOL].value and row[constants.ACTARRIVALCOL].value != " " and row[constants.ACTARRIVALCOL].value != "" and row[constants.ACTARRIVALCOL].value != "None"):
                    hold =True
                fillColor = "FFFFFF"
#                 
#                 if malPort:
#                     fillColor="FFFF00"
                if hold:
                    fillColor="FFC300"
#                 if hold and malPort:
#                     fillColor = "FF0000"
                
                
                for j in range(3):
                    row[j].fill = PatternFill(start_color=fillColor,
                    end_color=fillColor,
                    fill_type='solid')
        
        i+=1
        dims = {}
        for row in localContainers.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            localContainers.column_dimensions[col].width = value+3
        
        
        
def setupExcel(filepPath, containers):
      
    localContainersWb = load_workbook(filepPath, read_only=False, keep_vba=True)
    localContainers = localContainersWb["Current"]
    for cell in range(1, localContainers.max_column):
#         print(cell)
        val = localContainers[1][cell].value
        if val=="Container": 
            constants.CONTAINERCOL=cell
        elif val=="Load #": 
            constants.LOADCOL=cell
        elif val=="Size": 
            constants.SIZECOL=cell
        elif val=="L/E": 
            constants.LECOL=cell
        elif val=="Ship Point": 
            constants.SHIPPOINTCOL=cell
        elif val=="Current Position": 
            constants.CURRENTPOSCOL=cell
        elif val=="Last Reported": 
            constants.LASTREPORTEDCOL=cell
        elif val=="Equiment Status": 
            constants.STATUSCOL=cell
        elif val=="Cons Point": 
            constants.CONSPOINTCOL=cell
        elif val=="Grounded": 
            constants.GROUNDECOL=cell
        elif val=="Act Arrival": 
            constants.ACTARRIVALCOL=cell
        elif val=="Appointment Time": 
            constants.APPOINTMENTTIMECOL=cell
        elif val=="Holds": 
            constants.HOLDSCOL=cell
        elif val=="Last Free Day": 
            constants.LFDCOL=cell
        elif val=="Storage Guarantee": 
            constants.STORAGECOL=cell
        elif val=="Pieces": 
            constants.PIECESCOL=cell
        elif val=="Pickup Number": 
            constants.PUNUMBERCOL=cell
            
            
    for cell in range(2, localContainers.max_row+1):
        contNumber = localContainers[cell][int(constants.CONTAINERCOL)].value
#         print(contNumber)
        if contNumber != "" and contNumber != None:
            container = Container(contNumber.strip())
            containers.append(container)
    
    return localContainersWb, localContainers

# def copyRow(fromSheet, toSheet, indexCurrent, indexCompleted):
#     currentRow = fromSheet[indexCurrent]
#     completedRow = toSheet[indexCompleted]
#     for i in range(1, len(currentRow)):
#         new_cell=""
#         try:
#             new_cell = completedRow[i]
#         except:
#             new_cell = fromSheet.cell(row=indexCompleted, column=i)
#         cell=currentRow[i]
#         new_cell.value = currentRow[i].value
#         if cell.has_style:
#             new_cell.font = copy(cell.font)
#             new_cell.border = copy(cell.border)
#             new_cell.fill = copy(cell.fill)
#             new_cell.number_format = copy(cell.number_format)
#             new_cell.protection = copy(cell.protection)
#             new_cell.alignment = copy(cell.alignment)
    
# def moveCompleted(localContainersWb, filePath):
#     completed = localContainersWb["Completed"]
#     current = localContainersWb["Current"]
#     
# #     copyRow(current, completed, 1, 1)
#     destinationRow = completed.get_highest_row()
#     
#     for i in range(1, current.get_highest_row):
#         if "Load Out-Gate" in current.cell(row=i, column=7).value:
#             copyRow(current, completed, indexCurrent, indexCompleted) 
#     
#     
#     
#     dims = {}
#     for row in completed.rows:
#         for cell in row:
#             if cell.value:
#                 dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
#     for col, value in dims.items():
#         completed.column_dimensions[col].width = value+3
    
            
if __name__ == '__main__':
#     try:
    filePath = r"J:\LOCAL DEPARTMENT\Automation - DO NOT MOVE\Incoming Local Containers - CP.xlsm"
#     filePath = r"C:\Incoming Local Containers - CP.xlsm"
#     filePath = r"C:\Users\ssleep\Documents\Incoming Local Containers - CP.xlsm"
    containers = []
    
    localContainersWb, localContainers = setupExcel(filePath, containers)
    driver = setupCP()
    readContainerInfo(driver, containers)
    putInfoinExcel(containers, localContainers)
#     except:
#         print(exc_info())
#         sleep(60)
    saved = False
    while not saved:
        try:
            localContainersWb.save(filePath)
            saved = True
        except:
            top = Tk()
            L1 = Label()
            L1 = Label(top, text="Please close \"Incoming Local Containers - CP\" \nspreadsheet then hit \"OK\"")
            L1.config(font=("Courier", 16))
            L1.grid(row=0, column=0)
            top.lift()
            top.attributes('-topmost',True)
            top.after_idle(top.attributes,'-topmost',False)
              
            def callbackOK():
                top.destroy()
              
            MyButton5 = Button(top, text="OK", width=8, command=callbackOK)
            MyButton5.grid(row=1, column=0)
            MyButton5.config(font=("Courier", 16))
              
              
            w = 530 # width for the Tk root
            h = 100 # height for the Tk root
               
            # get screen width and height
            ws = top.winfo_screenwidth() # width of the screen
            hs = top.winfo_screenheight() # height of the screen
               
            # calculate x and y coordinates for the Tk root window
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
               
            # set the dimensions of the screen 
            # and where it is placed
            top.geometry('%dx%d+%d+%d' % (w, h, x, y))
           
            top.mainloop()
#     driver.quit()
#     driver.find_element_by_tag_name('html').send_keys(Keys.CONTROL, Keys.SHIFT, "w")
#     ActionChains(driver).send_keys(Keys.CONTROL, Keys.SHIFT, "w").perform()

#     sleep(1)
#     while True:
#         for p in process_iter():
#             if "geckodriver.exe" in p.name():
#                 p.kill()
#             break
    system('start excel.exe "'+ filePath + '"')
    
# pyinstaller "C:\Users\ssleep\workspace\CPTracer\Automator\__init__.py" --distpath "J:\Spencer\CPTracer" --noconsole -y
