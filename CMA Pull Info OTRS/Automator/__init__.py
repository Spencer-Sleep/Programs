from time import sleep
import sys
import atexit
def exit_hander():
    sleep(200)
atexit.register(exit_hander)
from pyotrs import Client
import warnings

from shutil import copyfile

import PyPDF2 
from PyPDF2.pdf import PdfFileReader
from _io import BytesIO, StringIO
from os import listdir
from os.path import isfile, join

from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfpage import PDFTextExtractionNotAllowed
from pdfminer.pdfinterp import PDFResourceManager
from pdfminer.pdfinterp import PDFPageInterpreter
from pdfminer.pdfdevice import PDFDevice
from pdfminer.layout import LAParams
from pdfminer.converter import PDFPageAggregator

import re

import pdfminer

# import pdftotext

# import textract

import base64
from openpyxl.reader.excel import load_workbook
import os

from openpyxl.workbook.workbook import Workbook
from openpyxl.styles.fonts import Font
from openpyxl.styles import colors
from datetime import timedelta
import datetime
# except:
#     print(sys.exc_info())
class Container(object):
    cnumber=""
    bookingNumber = ""
    WONumber = ""
    receivedTime=""
    address1=""
    address2=""
    address3=""
    vessel=""
#     eta=""
    cut=""
    
    cancelled=False

def extractText(pdf):
    try:
        pdf.save_to_dir(r"J:\Spencer\CMA Work Orders")
    except:pass
    
    name = r"J:\Spencer\CMA Work Orders\\"+pdf.Filename
    file = open(name, "rb")
#     print(ticket.articles[0].attachments[0].Content)
#     fileObj = StringIO()
#     fileObj.write(ticket.articles[0].attachments[0].Content)
#     fileData = base64.urlsafe_b64decode(pdf.encode('UTF-8'))
#     stream = BytesIO(fileData)


    parser = PDFParser(file)
    
    document = PDFDocument(parser)

    if not document.is_extractable:
        raise PDFTextExtractionNotAllowed
    
    # Create a PDF resource manager object that stores shared resources.
    rsrcmgr = PDFResourceManager()
    
    # BEGIN LAYOUT ANALYSIS
    # Set parameters for analysis.
    laparams = LAParams()
    
    # Create a PDF page aggregator object.
    device = PDFPageAggregator(rsrcmgr, laparams=laparams)
    
    # Create a PDF interpreter object.
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    
    text=""
    def parse_obj(text, lt_objs):
    
        # loop over the object list
        for obj in lt_objs:
    
            # if it's a textbox, print text and location
            if isinstance(obj, pdfminer.layout.LTTextBoxHorizontal):
                internalText=obj.get_text().replace('\n', '_')
                
                if(internalText):
                    text = text + internalText
            
#                 print (str(obj.bbox[0])+" "+ str(obj.bbox[1])+" "+ obj.get_text().replace('\n', '_'))
    
            # if it's a container, recurse
            elif isinstance(obj, pdfminer.layout.LTFigure):
                text = parse_obj(text, obj._objs)
        
        return text
    # loop over all pages in the document
    
    for page in PDFPage.create_pages(document):

        # read the page into a layout object
        interpreter.process_page(page)
        layout = device.get_result()
    
        # extract text from this object
        text = parse_obj(text, layout._objs)
#         if(thisText):
#             text = text + thisText
#     print(text)
    return text

def extractTextHapag(pdf):
    try:
        pdf.save_to_dir(r"J:\Spencer\CMA Work Orders")
    except:pass
    
    name = r"J:\Spencer\CMA Work Orders\\"+pdf.Filename
    pdfFileObj = open(name, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
    
    text = ''
    if pdfReader.isEncrypted:
        pdfReader.decrypt('')
    for i in range(pdfReader.getNumPages()):
        pageObj = pdfReader.getPage(i)
        text += pageObj.extractText()
    
    return text
    
def getRestOfLine(text, findString, instance=-1, ):
    if instance<1:
        startIndex = text.rfind(findString)
    elif instance >0:
        startIndex=0
        for _ in range(instance):
            oldStartIndex=startIndex
            startIndex = text[startIndex:].find(findString)
            if startIndex==-1:
                return ""
            startIndex+=+len(findString)+oldStartIndex
            
    if startIndex==-1:
        return ""
    if instance<0:
        startIndex+=len(findString)
    
    endIndex = text[startIndex:].find("_")+startIndex
#     print(text[startIndex:endIndex])
    return text[startIndex:endIndex]
    
    
def getContainer(text):
    container = Container()
    
    if "KCAN" in text:
        container.WONumber=getRestOfLine(text, "KCAN", 0)
        container.cut=getRestOfLine(text, "Restitution Date:_")
    elif "TCAN" in text:
        container.WONumber=getRestOfLine(text, "TCAN", 0)
        container.cut=getRestOfLine(text, "Available From:_")
    elif "KNAM" in text:
        container.WONumber=getRestOfLine(text, "KNAM", 0)
    
    voyage = getRestOfLine(text, "_Vessel: _")
    container.vessel=getRestOfLine(text, voyage+"_")+" "+voyage
    
    container.address1 = getRestOfLine(text, "_Address:_",1)
    container.address2 = getRestOfLine(text, "_Address:_",2)
    container.address3 = getRestOfLine(text, "_Address:_",3)
    
    container.cnumber=getRestOfLine(text, "_Container _")
    m = re.match("((?!TCAN)(?!KCAN)(?!KNAM)[A-Za-z]{4}[0-9]{7})", container.cnumber)
    if not m:
        m = re.search("((?!TCAN)(?!KCAN)(?!KNAM)[A-Za-z]{4}[0-9]{7})", text)
        if m:
            container.cnumber=m.group(0)
        else:
            numofcont=getRestOfLine(text, "_Containers: _")
            m = re.search("_[0-9]{2}[A-Za-z]{2}_", text[text.find("_Containers: _"):])
            container.cnumber = numofcont+"x"+ m.group(0)[1:-1]
        
#     if "Empty Repo" in text:
#         container.bookingNumber=getRestOfLine(text, "Empty Repo Ref: _SIPA: _")
#         container.vessel=""
    if "Booking Ref:  " in text:
        container.bookingNumber=getRestOfLine(text, "Booking Ref:  ")
    elif "B/L Ref:  " in text:
        container.bookingNumber=getRestOfLine(text, "B/L Ref:  ")
#     if "Customs Requirement" in container.cnumber:
#         print(text)
    return container
#     print(container)

def getContainerHapag(text):
    print(text)
    exit()
    
    containers=[]
    print(text)
    containerText = text[text.find("_W o r k   O r d e r   C o s t   S u m m a r y_"):]
    
#     m = re.findall("[A-Za-z]{4}  [0-9]{7}", containerText)
#     print(m)


    newContainer=Container()
    
    newContainer.cnumber=getRestOfLine(text, "Container:_",1)
    
    if newContainer.cnumber=="":
        newContainer.cnumber="Export"
        containers.append(newContainer)
    else:
        noContFound = False
        i=2
        while not noContFound:
            newContainer = Container()
            newContainer.cnumber=container
            containers.append(newContainer)
    
    
    
    
    
    
    
    
    
    
#     container = Container()
        
    
    for container in containers:
        container.address1 = getRestOfLine(text, "Pick up_")
        container.address2=getRestOfLine(text, "_To_")
        if "_Empty return to_" in text:
            container.address3=getRestOfLine(text, "_Empty return to_")
        
        findString="_Work Order: "
        startIndex =text.rfind(findString)+len(findString)
        endIndex = text[startIndex:].find(" ")+startIndex
        endIndex = min(text[startIndex:].find("_")+startIndex, endIndex)
        
        container.WONumber=text[startIndex:endIndex]
        
        if "_Del. Date_" in text:
            container.cut="Delivery Date: " + getRestOfLine(text, "_Del. Date_")
        elif "_Cutoff_" in text:
            container.cut="Cutoff: " + getRestOfLine(text, "_Cutoff_")
        else:
            container.cut="ETA: " + getRestOfLine(text, "_Arrival Date:_")
            
            
        fromLoc = text.find("_From")
        m = re.search("(-001_[0-9]{6}[ _])(.*?)_", text)
        if m:
            container.vessel=m.group(2)
#         else:
            if m.start()<fromLoc:
                startIndex=text.find(container.vessel)
                startIndex = text[startIndex:].find("_")+startIndex+1
                endIndex = text[startIndex:].find("_")+startIndex
                vesselCode = text[startIndex:endIndex]
                startIndex=endIndex+1
                endIndex = text[startIndex:].find("_")+startIndex
                container.vessel = text[startIndex:endIndex]+" "+vesselCode
        
        
        findString=container.vessel+"_"
        startIndex =text.find(findString)+len(findString)
        startIndex = text[startIndex:].find("_")+startIndex+1
        endIndex = text[startIndex:].find("_")+startIndex
        
        container.vessel += " " + text[startIndex:endIndex]
        
    return containers
    
    if "KCAN" in text:
        container.WONumber=getRestOfLine(text, "KCAN", 0)
        container.cut=getRestOfLine(text, "Restitution Date:_")
    elif "TCAN" in text:
        container.WONumber=getRestOfLine(text, "TCAN", 0)
        container.cut=getRestOfLine(text, "Available From:_")
    elif "KNAM" in text:
        container.WONumber=getRestOfLine(text, "KNAM", 0)
    
    voyage = getRestOfLine(text, "_Vessel: _")
    container.vessel=getRestOfLine(text, voyage+"_")+" "+voyage
    
    container.address1 = getRestOfLine(text, "_Address:_",1)
    container.address2 = getRestOfLine(text, "_Address:_",2)
    container.address3 = getRestOfLine(text, "_Address:_",3)
    
    container.cnumber=getRestOfLine(text, "_Container _")
    m = re.match("((?!TCAN)(?!KCAN)(?!KNAM)[A-Za-z]{4}[0-9]{7})", container.cnumber)
    if not m:
        m = re.search("((?!TCAN)(?!KCAN)(?!KNAM)[A-Za-z]{4}[0-9]{7})", text)
        if m:
            container.cnumber=m.group(0)
        else:
            numofcont=getRestOfLine(text, "_Containers: _")
            m = re.search("_[0-9]{2}[A-Za-z]{2}_", text[text.find("_Containers: _"):])
            container.cnumber = numofcont+"x"+ m.group(0)[1:-1]
        
#     if "Empty Repo" in text:
#         container.bookingNumber=getRestOfLine(text, "Empty Repo Ref: _SIPA: _")
#         container.vessel=""
    if "Booking Ref:  " in text:
        container.bookingNumber=getRestOfLine(text, "Booking Ref:  ")
    elif "B/L Ref:  " in text:
        container.bookingNumber=getRestOfLine(text, "B/L Ref:  ")
#     if "Customs Requirement" in container.cnumber:
#         print(text)
    return container

def putContainerInSheet(containers, sheetLocation, listBook, listSheet, lastFileNumber, code):
    containerDict = dict()
    i=1
    for row in listSheet.rows:
        try:
            containerDict[row[3].value]=i
            i+=1
        except:
            pass
#     i=1
    containers.reverse()
    for container in containers:
        lastRow = listSheet.max_row+1
        values = [container.cnumber, container.WONumber, container.bookingNumber, container.vessel, container.address1, container.address2]
        
        if code=="TCAN":
            values.append(container.address3)
        if code!="KNAM":
            values.append(container.cut)
        values.append(container.receivedTime)
        if code=="TCAN":
            values.append("=HYPERLINK(\"J:\\LOCAL DEPARTMENT\\CMA WOs\\"+container.WONumber+"-"+container.cnumber+".pdf\", \"Work Order\")")
        if container in containerDict.keys():
            lastRow=containerDict[container]
            if container.cancelled:                
                listSheet.cell(lastRow,len(values)+1).value = "Cancelled"
            else:
                listSheet.cell(lastRow,len(values)+1).value = "Amended"
            
#         print(lastRow)
        i=1
        for value in values:
            listSheet.cell(lastRow,i).value = value
            if "HYPER" in value:
                listSheet.cell(lastRow,i).font = Font(u='single', color=colors.BLUE)
#         listSheet.cell(lastRow,1).value = container.WONumber
#         listSheet.cell(lastRow,2).value = container.bookingNumber
#         listSheet.cell(lastRow,3).value = container.cnumber
#         listSheet.cell(lastRow,4).value = container.vessel
#         listSheet.cell(lastRow,5).value = container.address1
#         listSheet.cell(lastRow,6).value = container.address2
#         listSheet.cell(lastRow,7).value = container.address3
#         listSheet.cell(lastRow,8).value = container.cut
#         listSheet.cell(lastRow,9).value = container.receivedTime
            i+=1
#             
        if container.cancelled:
            listSheet.cell(lastRow,i).value = "Cancelled"
#         i+=1
    if len(containers) != 0:
        listSheet.cell(1,i+2).value = datetime.datetime.utcnow()
    i=lastFileNumber
#     print(i)
    deleted = True
    while deleted:
        try:
            os.remove(sheetLocation+"List"+str(i)+".xlsx")
            i=i-1
            if i<1:
                deleted=False
                i=1
        except:
            i+=1
            deleted=False
            
    dims = {}
    for row in listSheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
    for col, value in dims.items():
        listSheet.column_dimensions[col].width = value+2
    if code=="TCAN":
        listSheet.column_dimensions["J"].width=13
    saved = False
    while not saved:
        try:
            listBook.save(sheetLocation+"List"+str(i)+".xlsx")
            saved = True
        except:
            i+=1

def putContainerInSheetHapag(containers, sheetLocation, listBook, listSheet, lastFileNumber):
    print(sheetLocation)
    containerDict = dict()
    i=1
    print(containers)
    for row in listSheet.rows:
        try:
            containerDict[row[3].value]=i
            i+=1
        except:
            pass
#     i=1
#     containers.reverse()
    for container in containers:
        lastRow = listSheet.max_row+1
        values = [container.cnumber, container.WONumber, container.bookingNumber, container.vessel, container.address1, container.address2, container.address3, container.cut, container.receivedTime]
        
        values.append("=HYPERLINK(\"J:\\LOCAL DEPARTMENT\\Hapag WO\\"+container.WONumber+"-"+container.cnumber+".pdf\", \"Work Order\")")
        if container in containerDict.keys():
            lastRow=containerDict[container]
            
#         print(lastRow)
        i=1
        for value in values:
            listSheet.cell(lastRow,i).value = value
            if "HYPER" in value:
                listSheet.cell(lastRow,i).font = Font(u='single', color=colors.BLUE)
            i+=1
#             
#         i+=1
    if len(containers) != 0:
        listSheet.cell(1,i+2).value = datetime.datetime.utcnow()
    i=lastFileNumber
#     print(i)
    deleted = True
    while deleted:
        try:
            os.remove(sheetLocation+"List"+str(i)+".xlsx")
            i=i-1
            if i<1:
                deleted=False
                i=1
        except:
            i+=1
            deleted=False
            
    dims = {}
    for row in listSheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
    for col, value in dims.items():
        listSheet.column_dimensions[col].width = value+2
    listSheet.column_dimensions["J"].width=13
    saved = False
    while not saved:
        try:
            listBook.save(sheetLocation+"List"+str(i)+".xlsx")
            saved = True
        except:
            i+=1

def fetchCMAWOInfo(client, codes, sheetLocations):
    
    imports=[]
    exports=[]
    local=[]
    listBooks=[]
    listSheets=[]
    lastFileNumbers=[]
    for i in range(3):
        code=codes[i]
        sheetLocation=sheetLocations[i]
        onlyfiles = [f for f in listdir(sheetLocation) if isfile(join(sheetLocation, f))]
        lastFile = ""
        lastFileNumber = 0
        
        for file in onlyfiles:
            if not file[0]=="~":
                fileNumber = int(file[4:file.find(".")])
                if fileNumber>lastFileNumber:
                    lastFile=file
                    lastFileNumber=fileNumber
        
                
        lastFileNumbers.append(lastFileNumber)
        try:
            listBook = load_workbook(sheetLocation+lastFile)
            listSheet = listBook.active
        except:
            listBook=Workbook()
            listSheet = listBook.active
            
            headers = ["Container Number", "WO Number", "Booking Number", "Vessel", "Address 1", "Address 2"]
            
            if code=="TCAN":
                headers = headers + ["Address 3", "Available Date"]
            if code=="KCAN":
                headers.append("Cut")
            
            headers.append("Date Received")
            if code=="TCAN":
                headers.append("WO Link")
            
            
            i=1
            for header in headers:
                listSheet.cell(1,i).value = header
                i+=1
                
            listSheet.cell(1,i+1).value = "Last Checked:"
            listSheet.cell(1,i+2).value = None
            
            
    #         listSheet.cell(1,1).value = "WO Number"
    #         listSheet.cell(1,2).value = "Booking Number"
    #         listSheet.cell(1,3).value = "Container Number"
    #         listSheet.cell(1,4).value = "Vessel"
    #         listSheet.cell(1,5).value = "Address 1"
    #         listSheet.cell(1,6).value = "Address 2"
    #         listSheet.cell(1,7).value = "Address 3"
    #         listSheet.cell(1,8).value = "Cut"
    #         listSheet.cell(1,9).value = "Date Received"
            
            
            
    #         listSheet.cell(1,11).value = "Last Checked:"
        
        
        
    #     tickets = client.ticket_search(Title="Transport Order: TCAN1817410")
        
    #     tickets = client.ticket_search(Title="Transport Order: " + code+"*")
    #     if False:
    #     tickets=""
        listBooks.append(listBook)
        listSheets.append(listSheet)
        lastCheckLoc = 10
        if code=="KCAN":
            lastCheckLoc+=1
        elif code=="TCAN":
            lastCheckLoc+=3
    #     print("Transport Order: " + code+"*")
        if listSheet.cell(1,lastCheckLoc).value=="" or listSheet.cell(1,lastCheckLoc).value==None:
#             if code=="TCAN":
#                 tickets = client.ticket_search(Title="Transport Order: TCAN1815669")
#             else:
            tickets=client.ticket_search(Title="Transport Order*" + code+"*")
        else:
            tickets=client.ticket_search(Title="Transport Order*" + code+"*", TicketCreateTimeNewerDate=listSheet.cell(1,lastCheckLoc).value)
        containers=[]
        containerNumbers=[]
        for ticket in tickets:
            if not ticket:
                continue
            ticket = (client.ticket_get_by_id(ticket, True, True))
        #     ticket.articles[0].to_dct()
    #         print(ticket.articles[0].field_get("Subject"))
            
        #     print(ticket.articles[0].attachments[0])
        
            #     pdf = ticket.articles[0].attachments[0].Content
            for article in ticket.articles:
                for attachment in article.attachments:
                    pdf = attachment
                    if ".pdf" in attachment.Filename:
                        text = extractText(pdf)
                        if not "Empty Repo" in text and text!="":
                            container = getContainer(text)
                            print(container.WONumber)
                            if "TCAN" in ticket.articles[0].field_get("Subject"):
                                try:
                                    copyfile(r"J:\Spencer\CMA Work Orders\\"+pdf.Filename, "J:\LOCAL DEPARTMENT\CMA WOs\\"+container.WONumber+"-"+container.cnumber+".pdf")
                                except:
                                    pass
                            
                            if not container.cnumber in containerNumbers:
                                containerNumbers.append(container.cnumber)
                #                 print(container.cnumber)
                                container.cancelled="Cancellation" in ticket.articles[0].field_get("Subject")
                                container.receivedTime=str(datetime.datetime.strptime(ticket.articles[0].field_get("CreateTime"), "%Y-%m-%d %H:%M:%S")-timedelta(hours=4))
                                containers.append(container)
                    else:
                        print(attachment.Filename)
        if i==0:
            exports = list(containers)
        elif i==1:
            local = list(containers)
        elif i==2:
            imports = list(containers)
        
    expToSwitch=[]
    impToSwitch=[]
        
    for container in exports:
        if "Seaport Intermodal" in container.address2:
            expToSwitch.append(container)
    for container in imports:
        if "Seaport Intermodal" in container.address1:
            impToSwitch.append(container)
            
    for container in expToSwitch:
        imports.append(container)
        exports.remove(container)
    for container in impToSwitch:
        imports.remove(container)
        exports.append(container)
    
    
    putContainerInSheet(exports, sheetLocations[0], listBooks[0], listSheets[0], lastFileNumbers[0],codes[0])
    putContainerInSheet(local, sheetLocations[1], listBooks[1], listSheets[1], lastFileNumbers[1],codes[1])
    putContainerInSheet(imports, sheetLocations[2], listBooks[2], listSheets[2], lastFileNumbers[2],codes[2])

def fetchHAPAG():
    hapagLocation = r"J:\LOCAL DEPARTMENT\Hapag WO"
    hapagSheetLocation = hapagLocation+" Sheets\\"
    hapagLocation+="\\"
    onlyfiles = [f for f in listdir(hapagSheetLocation) if isfile(join(hapagSheetLocation, f))]
    lastFile = ""
    lastFileNumber = 0
    
    for file in onlyfiles:
        if not file[0]=="~":
            fileNumber = int(file[4:file.find(".")])
            if fileNumber>lastFileNumber:
                lastFile=file
                lastFileNumber=fileNumber
    
            
    try:
        listBook = load_workbook(hapagSheetLocation+lastFile)
        listSheet = listBook.active
    except:
        listBook=Workbook()
        listSheet = listBook.active
        
        headers = ["Container Number", "WO Number", "Booking Number", "Vessel", "Address 1", "Address 2", "Address 3", "Relevant Date", "Date Received", "WO Link"]
        
        
        i=1
        for header in headers:
            listSheet.cell(1,i).value = header
            i+=1
            
        listSheet.cell(1,i+1).value = "Last Checked:"
        listSheet.cell(1,i+2).value = None
            
    if listSheet.cell(1,13).value=="" or listSheet.cell(1,13).value==None:
#         tickets=client.ticket_search(Title="WOSD0001*")
        tickets=client.ticket_search(Title="WOSD0001*", TicketChangeTimeNewerDate="2018-10-16 00:00:00")
    else:
        tickets=client.ticket_search(Title="WOSD0001*", TicketChangeTimeNewerDate=listSheet.cell(1,13).value)
    
#     tickets=[tickets[:4]]
#     print(tickets)
    
    containers=[]
    woNumbers=[]
    for ticket in tickets:
        if not ticket:
            continue
        ticket = (client.ticket_get_by_id(ticket, True, True))
        for article in reversed(ticket.articles):
            for attachment in article.attachments:
                pdf = attachment
                if ".pdf" in attachment.Filename and "WOSD0001" in attachment.Filename:
#                     print(attachment.Filename)
                    text = extractTextHapag(pdf)
#                     if not "Empty Repo" in text and text!="":
                    containersParsed = getContainerHapag(text)
#                         if "TCAN" in ticket.articles[0].field_get("Subject"):
                    try:
                        if len(containersParsed)==1:
                            copyfile(r"J:\Spencer\CMA Work Orders\\"+pdf.Filename, "J:\LOCAL DEPARTMENT\Hapag WO\\"+containersParsed[0].WONumber+"-"+containersParsed[0].cnumber+".pdf")
                        else:
                            contString=""
                            for container in containersParsed:
                                contString+=container.cnumber+"-"
                            contString=contString[:-1]
                            copyfile(r"J:\Spencer\CMA Work Orders\\"+pdf.Filename, "J:\LOCAL DEPARTMENT\Hapag WO\\"+containersParsed[0].WONumber+"-"+contString+".pdf")
                    except:
                        print(sys.exc_info())
                        pass
                    if not containersParsed[0].WONumber in woNumbers:
                        print(containersParsed[0].WONumber)
                        for container in containersParsed:
                            woNumbers.append(container.WONumber)
            #                 print(container.cnumber)
#                                 container.cancelled="Cancellation" in ticket.articles[0].field_get("Subject")
                            container.receivedTime=str(datetime.datetime.strptime(ticket.articles[0].field_get("ChangeTime"), "%Y-%m-%d %H:%M:%S")-timedelta(hours=4))
                            containers.append(container)
#                 else:
#                     print(attachment.Filename)
    putContainerInSheetHapag(containers, hapagSheetLocation, listBook, listSheet, lastFileNumber)
#     exit()
    
    
if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    
    client = Client("https://core.seaportint.com/", "testadmin", "testpass")
    a = client.session_create()
    if(a):
        print("Connected to OTRS as Testadmin")
    while True:
        print("Fetching")
        Codes=["KCAN", "TCAN", "KNAM"]
    #     Codes=["TCAN"]
        sheetLocations=[r"J:\ANTONIO-Export Work\CMA WO Sheets\\", r"J:\LOCAL DEPARTMENT\CMA WO Sheets\\", r"J:\IMPORTS\CMA WO Sheets\\"]
#         sheetLocations=[r"C:\Users\ssleep\Documents\CMA WO\\", r"J:\LOCAL DEPARTMENT\CMA WO Sheets\\", r"J:\IMPORTS\CMA WO Sheets\\"]
#         fetchCMAWOInfo(client, Codes, sheetLocations)
        fetchHAPAG()
        print("Done")
        sleep(600)
    
    
#pyinstaller "C:\Users\ssleep\workspace\CMA Pull Info OTRS\Automator\__init__.py" --distpath "J:\Spencer\OTRS Daemon" -y
    