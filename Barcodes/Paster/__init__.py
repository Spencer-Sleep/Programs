from openpyxl import load_workbook
import pythoncom
import pyHook
from win32api import GetKeyState
from win32con import VK_CAPITAL
from pyautogui import click, press, hotkey, typewrite, keyDown, keyUp, position
from time import sleep
from pywinauto.win32functions import SendInput
import pyperclip
import ctypes
import time
from past.builtins.misc import unichr
import unicodedata
from HelperFunctions import done
import pyautogui
from openpyxl.workbook.workbook import Workbook
from sspicon import MsV1_0Lm20ChallengeRequest
SendInput = ctypes.windll.user32.SendInput

PUL = ctypes.POINTER(ctypes.c_ulong)
class KeyBdInput(ctypes.Structure):
    _fields_ = [("wVk", ctypes.c_ushort),
                ("wScan", ctypes.c_ushort),
                ("dwFlags", ctypes.c_ulong),
                ("time", ctypes.c_ulong),
                ("dwExtraInfo", PUL)]

class HardwareInput(ctypes.Structure):
    _fields_ = [("uMsg", ctypes.c_ulong),
                ("wParamL", ctypes.c_short),
                ("wParamH", ctypes.c_ushort)]

class MouseInput(ctypes.Structure):
    _fields_ = [("dx", ctypes.c_long),
                ("dy", ctypes.c_long),
                ("mouseData", ctypes.c_ulong),
                ("dwFlags", ctypes.c_ulong),
                ("time",ctypes.c_ulong),
                ("dwExtraInfo", PUL)]

class Input_I(ctypes.Union):
    _fields_ = [("ki", KeyBdInput),
                 ("mi", MouseInput),
                 ("hi", HardwareInput)]

class Input(ctypes.Structure):
    _fields_ = [("type", ctypes.c_ulong),
                ("ii", Input_I)]
    
KEYEVENTF_UNICODE = 0x0004 
KEYEVENTF_KEYUP = 0x0002

def PressKey(KeyUnicode):

    extra = ctypes.c_ulong(0)
    ii_ = Input_I()
    ii_.ki = KeyBdInput( 0, KeyUnicode, KEYEVENTF_UNICODE, 0, ctypes.pointer(extra) )
    x = Input( ctypes.c_ulong(1), ii_ )
    ctypes.windll.user32.SendInput(1, ctypes.pointer(x), ctypes.sizeof(x))

def ReleaseKey(KeyUnicode):

    extra = ctypes.c_ulong(0)
    ii_ = Input_I()
    ii_.ki = KeyBdInput( 0, KeyUnicode, KEYEVENTF_UNICODE|KEYEVENTF_KEYUP, 0, ctypes.pointer(extra) )
    x = Input( ctypes.c_ulong(1), ii_ )
    ctypes.windll.user32.SendInput(1, ctypes.pointer(x), ctypes.sizeof(x))

if __name__ == '__main__':
#     routingBook = load_workbook(r"C:\Users\ssleep\Documents\Contacts to combine\combineFiltered.xlsx")
# #     routingBook = load_workbook(r"C:\Users\ssleep\Documents\Contacts to combine\CustomerIDs.xlsx")
#     mainSheet = routingBook.active
#     
#     for row in mainSheet.rows:
#         row[1].value = row[1].value.strip()
#     
#     routingBook.save(r"C:\Users\ssleep\Documents\Contacts to combine\combineFiltered.xlsx")
#     exit()
#     
#     routingBook = load_workbook(r"C:\Users\ssleep\Documents\Contacts to combine\combine.xlsx")
# #     routingBook = load_workbook(r"C:\Users\ssleep\Documents\Contacts to combine\CustomerIDs.xlsx")
#     mainSheet = routingBook.active
#     
#     destBook = Workbook()
#     ds = destBook.active    
#     
#     emails = []
#     col = 1
#     for row in mainSheet.rows:
#         if not row[1].value.lower() in emails:
#             if "ca.nestle.com" in row[1].value:
#                 name = row[1].value.lower()
#             else:
#                 name = row[0].value.lower()
#             if name[0]=="'":
#                 name=name[1:]
#             if name[-1]=="'":
#                 name=name[:-1]
#             name = name.replace("("," ")
#             name = name.replace(")","")
#             if "@" in name:
#                 name = name[:name.find("@")]
#             if "," in name:
#                 i=0
#                 firstname = ""
#                 for nameX in name.split(","):
#                     if i==0:
#                         lastname = nameX.strip()
#                     else:
#                         firstname = firstname + nameX.strip()
#                     i+=1
#             elif " " in name:
#                 i=0
#                 lastname = ""
#                 for nameX in name.split(" "):
#                     if i==0:
#                         firstname = nameX.strip()
#                     else:
#                         lastname = lastname + nameX.strip()
#                     i+=1
#             elif "." in row[0].value:
#                 i=0
#                 lastname = ""
#                 for nameX in name.split("."):
#                     if i==0:
#                         firstname = nameX.strip()
#                     else:
#                         lastname = lastname + nameX.strip()
#                     i+=1
#             else:
#                 firstname = name
#                 lastname = " "
#             if lastname=="":
#                 lastname=" "
#             while lastname[0]=="-":
#                 lastname=lastname[1:]
#             if lastname[0]=="|":
#                 lastname=lastname[1:]
#             ds.cell(col, 1).value = firstname
#             ds.cell(col, 2).value = lastname
# #             ds.cell(col, 1).value = row[0].value.lower()
#             ds.cell(col, 3).value = row[1].value.lower()
#             ds.cell(col, 4).value = row[1].value.lower()
#             index = row[2].value.rfind(".")
#             ds.cell(col, 5).value = row[2].value.lower()[:index]
#             ds.cell(col, 6).value = "valid"
#             col+=1
#             emails.append(row[1].value.lower())
# #         if row[0].value and not row[0].value.lower() in emails:
# # #             ds.cell(col, 1).value = row[0].value.lower()
# #             index = row[0].value.rfind(".")
# #             ds.cell(col, 1).value = row[0].value[0].lower()+row[0].value[1:index].lower()
# #             ds.cell(col, 2).value = row[0].value[0].lower()+row[0].value[1:index].lower()
# #             ds.cell(col, 3).value= "valid"
# #             col+=1
# #             emails.append(row[0].value.lower())
#     destBook.save(r"C:\Users\ssleep\Documents\Contacts to combine\combineFiltered.xlsx")
# #     destBook.save(r"C:\Users\ssleep\Documents\Contacts to combine\CustomerIDsFiltered.xlsx")
#     
#     exit()
    
    
    
    print("CREATE THE FIRST BARCODE AS NORMAL AND COPY IT.\n")
    print("FOR SUBSEQUENT BARCODES, PASTE THE FIRST BARCODE")
    print("THEN HOVER THE CURSOR OVER IT AND PRESS \"CAPS LOCK\"")
#     print(chr(0xC3))
#     PressKey(0xAE)
#     ReleaseKey(0xAE)
#     exit()
    
    codeBook = load_workbook(r"C:\Automation\barcode-generator-excel.xlsm", data_only=True)
    
    codeSheet = codeBook.active
    
    for cell in next(codeSheet.rows):
        if cell.value and "Item Number" in cell.value:
            itemCol = cell.col_idx - 1
        if cell.value and "Barcode" in cell.value:
            barcodeCol = cell.col_idx - 1

    pars = []
    
    first = False
    for row in codeSheet.rows:
#         print(str(row[itemCol].value)[0:4])
        if str(row[itemCol].value)[0:4]=="20C0":
            pars.append((row[itemCol].value, row[barcodeCol].value))
    try:
        pars.pop(0)
    except:
        print("NO VALID PARS FOUND")
        sleep(100)
        exit()
    print("\n")
    print("PARS PASTED:")
#     print("\n")
    while len(pars)>0:
        if GetKeyState(27) < 0:
                exit()
                
        while not GetKeyState(20)<0:
            True
        
        if GetKeyState(27) < 0:
            exit()
        
            
        click(clicks=2)
        
        press("up")
        press("end")
        press("backspace", 20)
        
        if GetKeyState(VK_CAPITAL):
            press("capslock")
        
        parsX = pars.pop(0)
        typewrite(str(parsX[0]))
        
        press("right", 2)
        press("delete", 20)
        press("backspace")
        
        
        for i in range(len(str(parsX[1]))):
            s = str(parsX[1])[i]
            unicodedata.name(s)
            PressKey(ord(s))
            ReleaseKey(ord(s))

        print(parsX[0])            
        sleep(1)
    
    done()
    
    ####################
    #REPO THING#########
    ####################
    
#     print("CREATE THE FIRST BARCODE AS NORMAL AND COPY IT.\n")
#     print("FOR SUBSEQUENT BARCODES, PASTE THE FIRST BARCODE")
#     print("THEN HOVER THE CURSOR OVER IT AND PRESS \"CAPS LOCK\"")
# #     print(chr(0xC3))
# #     PressKey(0xAE)
# #     ReleaseKey(0xAE)
# #     exit()
#     
#     codeBook = load_workbook(r"C:\Automation\barcode-generator-excel.xlsm", data_only=True)
#     
#     codeSheet = codeBook.active
#     
#     for cell in next(codeSheet.rows):
#         if cell.value and "Item Number" in cell.value:
#             itemCol = cell.col_idx - 1
#         if cell.value and "Barcode" in cell.value:
#             barcodeCol = cell.col_idx - 1
# 
#     pars = []
#     
#     first = False
#     for row in codeSheet.rows:
# #         print(str(row[itemCol].value)[0:4])
# #         if str(row[itemCol].value)[0:4]=="20C0":
# 
# 
#         if row[0].row>6 and row[itemCol].value and row[itemCol].value!="None":
#             pair = row[itemCol].value.split("_x000D_\n")
#             print(pair)
#             pars.append(pair[0])
#             pars.append(pair[1])
#             
#             
# #             pars.append(row[itemCol].value)
# #             pars.append(row[itemCol].value)
#             
# #     try:
# #         pars.pop(0)
# #     except:
# #         print("NO VALID PARS FOUND")
# #         sleep(100)
# #         exit()
#     print("\n")
#     print("PARS PASTED:")
# #     print("\n")
#     while len(pars)>0:
#         if GetKeyState(27) < 0:
#                 exit()
#                 
#         while not GetKeyState(20)<0:
#             True
#         
#         if GetKeyState(27) < 0:
#             exit()
#         
#             
# #         click(clicks=1)
#         
# #         press("up")
# #         press("end")
# #         press("backspace", 20)
#         
#         if GetKeyState(VK_CAPITAL):
#             press("capslock")
#         
#         parsX = pars.pop(0)
#         
#         typewrite(str(parsX))
#          
# #         press("tab")
# #         typewrite(str(parsX))
# #         press("tab")
# #         typewrite(str(parsX))
#         
#         
# #         press("right", 2)
# #         press("delete", 20)
# #         press("backspace")
#         
#         
# #         for i in range(len(str(parsX[1]))):
# #             s = str(parsX[1])[i]
# #             unicodedata.name(s)
# #             PressKey(ord(s))
# #             ReleaseKey(ord(s))
# 
# #         print(parsX[0])            
#         sleep(1)
#     
#     done()
#     pyinstaller "C:\Users\ssleep\workspace\Barcodes\Paster\__init__.py" --distpath "J:\Spencer\Barcode Paster" -y