from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver import Firefox
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

from os import devnull
from os import system
# from win32com.client.gencache import EnsureDispatch #@UnresolvedImport
# from win32com.client import Dispatch #@UnresolvedImport
from openpyxl import load_workbook
from openpyxl.styles.colors import RGB
from openpyxl.styles import PatternFill
from copy import copy

from sys import path
from sys import exc_info
from time import sleep

from psutil import process_iter #@UnresolvedImport

from selenium.webdriver.common.action_chains import ActionChains

from tkinter import Button, Tk, Label, Entry

import re
from HelperFunctions import popUpOK
import sys
from _datetime import datetime
from pycparser.c_ast import Switch

# from enum.IntFlag import  



class constants:
    CONTAINERCOL = ''
    PICKUPCOL = ""
    STATUSCOL=""
    DOCUMENTSCOL=''
    CUSTOMSCOL=""
    STEAMSHIPCOL=""
    OTHERCOL=""
    ETACOL=""
    STORAGEDATECOL=""
    STORAGEAMOUNTCOL=""
    RVNUMBERCOL=""
    RVTIMECOL=""
    APPOINTMENTTIMECOL=""
    PBCOL=""
    SIZECOL=""


class Container:
    def __init__(self, number = ""):
        self.number = number.upper()
        self.pickupNumber = ""
        self.eta = ""
        self.status = ""
        self.documents = ""
        self.customs = ""
        self.steamship = ""
        self.storageDate =""
        self.storageAmount =""
        self.other= ""
        self.size=""
        self.rvtime=""
        self.rvnumber=""
        self.checkDigit=""

def setupCn():
    fp = FirefoxProfile();
    fp.set_preference("webdriver.load.strategy", "unstable");
#     fp.set_preference("XRE_NO_WINDOWS_CRASH_DIALOG=1")
     
    driver = Firefox(firefox_profile=fp, log_path=devnull)
    driver.get("http://cn.ca/")
#     driver.set_window_position(1920, 0)
#     sleep(30)
    driver.maximize_window()
    
    driver.implicitly_wait(40)
    
    f=open(r"J:\LOCAL DEPARTMENT\Automation - DO NOT MOVE\CN Login.txt", 'r')
#     f=open(r"C:\Automation\CN Login.txt", 'r')
    read = f.readline()
    m = re.search("username: *", read)
    username = read[m.end():].rstrip()
    read = f.readline()
    m = re.search("password: *", read)
    password = read[m.end():].rstrip()
    f.close()    
    
    driver.find_element_by_class_name("lbl").click()
    
    driver.find_element_by_id("login_usernameNew").send_keys(username)
    driver.find_element_by_id("login_passwordNew").send_keys(password)
    driver.find_element_by_id("loginform_enterbutton").click()
    
    driver.implicitly_wait(10)
    
    try:
        driver.find_element_by_css_selector("fa-icon[class='fa-fw ng-fa-icon'").click()
    except:
        print("no")
        pass
    
    driver.implicitly_wait(100)
    
    return driver
        
def readContainerInfo(driver, containers):
    driver.implicitly_wait(500)
#     sleep(30)
#     driver.switch_to.frame("menuHeader")
#     driver.find_element_by_id("id18").click()
    driver.implicitly_wait(1)
    try:
        driver.find_element_by_css_selector("ci-tools-standalone-menu[class='ci-recent-tools-container ng-tns-c5-3 ng-star-inserted']").click()
        driver.find_element_by_css_selector("a[href='#/tools/intermodal-shipment-status']").click()
    except:
        pass

    driver.switch_to_frame(driver.find_element_by_css_selector("iframe[name='ci-tools-frame']"))
    driver.switch_to_frame(driver.find_element_by_name("main"))
    
#     try:
#         driver.find_element_by_id("id18").click()
#     except:
#         driver.implicitly_wait(600)
# #         driver.find_element_by_id("tools").click()
#         driver.find_element_by_id("tools").click()
#         driver.find_element_by_id("tools").click()
#         driver.find_element_by_id("tools").click()
# #         sleep(5)
# #         driver.find_element_by_class_name("tools selected").click()
#         driver.switch_to_default_content()
#         driver.switch_to_frame("content1")
# #         sleep(500)
# #         driver.find_element_by_css_selector(r'a[href^="top.frames[0].openTab(\'id41\');"]').click()
# #         print(driver.find_element_by_css_selector('a[onclick*="top.frames[0].openTab(\'id41\');"]').text)
#         driver.find_element_by_css_selector('a[onclick*="top.frames[0].openTab(\'id18\');"]').click()
#         driver.switch_to_default_content()
#         driver.switch_to.frame("menuHeader")
#         driver.find_element_by_id("id18").click()
# #     driver.find_element_by_css_selector('a[id*="id18"][class*="label"]').click()
#     driver.switch_to_default_content()
#     driver.switch_to_frame(driver.find_element_by_css_selector("frame[name='content1']"))
#     i = 2
#     found = False
#     driver.implicitly_wait(0)
#     while not found:
#         try:
#             driver.switch_to_default_content()
#             driver.switch_to_frame(driver.find_element_by_css_selector("frame[name='content" + str(i) + "']"))
#             driver.switch_to_frame(driver.find_element_by_css_selector("frame[src*='CFF_ImdShipmentStatus']"))
#             found = True
#         except:
#             if i<30:
#                 i+=1
#             else:
#                 i=2
    driver.implicitly_wait(600)
#     driver.switch_to.frame("content3")
#     driver.switch_to.frame("main")
#     driver.switch_to_frame(driver.find_element_by_css_selector(""))
    
    containerText = driver.find_element_by_id("cars")
    for container in containers:
        containerText.send_keys(container.number + "\n")
        
    driver.find_element_by_id("responsetype-browser").click()
    driver.find_element_by_id("submit").click()
    
    i = 0
#     driver.find_element_by_css_selector("tbody")
    
    rows = driver.find_element_by_css_selector("table[class='data aligntop aligncenter']>tbody").find_elements_by_css_selector("tr")
    for row in rows:
        if i != 0:
            cells = row.find_elements_by_css_selector("td")
            container = ""
            matchString = cells[0].text
            while matchString[len(matchString)-1]==" ":
                matchString = matchString[:len(matchString)-2]
            m = re.compile(matchString[:4] + " ?0*" + matchString[5:], re.RegexFlag.IGNORECASE)
            for Xcontainer in containers:
#                 if Xcontainer.number[:4]+" "+Xcontainer.number[4:10] in cells[0].text:
                if m.search(Xcontainer.number):
                    container=Xcontainer
                    break
            j = 0
            if container != "":
                for cell in cells:
                    if j==1:
                        container.status = cell.text
                    elif j==2:
                        container.size = cell.text
                    elif j==3:
                        container.documents = cell.text
                    elif j==4:
                        container.customs = cell.text
                    elif j==5:
                        container.steamship = cell.text
                    elif j==6:
                        container.other = cell.text
                    elif j==7:
                        container.eta = cell.text
                    elif j==8:
                        container.storageDate = cell.text
                    elif j==9:
                        container.storageAmount = cell.text
                    j+=1
#                 else:
#                     print(container)
#                 print(cell.get_attribute("value"))
#                 print(cell.tag_name)
        i+=1
    try:
        driver.switch_to_default_content()
        driver.find_element_by_css_selector("ci-tools-standalone-menu[class='ci-recent-tools-container ng-tns-c5-3 ng-star-inserted']").click()
        driver.find_element_by_css_selector("a[href='#/tools/gate-appointment-inquiry']").click()
    except:
        pass

    driver.switch_to_frame(driver.find_element_by_css_selector("iframe[name='ci-tools-frame']"))
#     sleep(600)    
#     driver.switch_to_default_content()
#     driver.switch_to.frame("menuHeader")
# #     driver.find_element_by_id("id41").click()
#     driver.implicitly_wait(1)
#     try:
#         driver.find_element_by_id("id41").click()
#     except:
#         driver.implicitly_wait(600)
# #         driver.find_element_by_id("tools").click()
#         driver.find_element_by_id("tools").click()
#         driver.find_element_by_id("tools").click()
#         driver.find_element_by_id("tools").click()
# #         sleep(5)
# #         driver.find_element_by_class_name("tools selected").click()
#         driver.switch_to_default_content()
#         driver.switch_to_frame("content1")
# #         sleep(500)
# #         driver.find_element_by_css_selector(r'a[href^="top.frames[0].openTab(\'id41\');"]').click()
# #         print(driver.find_element_by_css_selector('a[onclick*="top.frames[0].openTab(\'id41\');"]').text)
#         driver.find_element_by_css_selector('a[onclick*="top.frames[0].openTab(\'id41\');"]').click()
#         driver.switch_to_default_content()
#         driver.switch_to.frame("menuHeader")
#         driver.find_element_by_id("id41").click()
# 
# 
#     driver.switch_to_default_content()
#     driver.switch_to_frame(driver.find_element_by_css_selector("frame[name='content1']"))
#     i = 2
#     found = False
#     driver.implicitly_wait(0)
#     while not found:
#         try:
#             driver.switch_to_default_content()
#             driver.switch_to_frame(driver.find_element_by_css_selector("frame[name='content" + str(i) + "']"))
#             driver.find_element_by_css_selector("form[action='AppointmentQuery']")
#             found = True
#         except:
#             if i<30:
#                 i+=1
#             else:
#                 i=2
    driver.implicitly_wait(600)
    for k in range(int(len(containers)/20)+1):
        driver.find_element_by_css_selector("input[value='equipment']").click()
        containerText = driver.find_element_by_name("ids")
        m = k*20
        while m < (k+1)*20:
            if m<len(containers):
                noLeadingZeroes = containers[m].number
                while noLeadingZeroes[4]=="0":
                    noLeadingZeroes = noLeadingZeroes[0:4] + noLeadingZeroes[5:]
                containerText.send_keys(noLeadingZeroes + "\n")
            m+=1
            
        driver.find_element_by_id("btn_23").click()
        i=0
        table = driver.find_element_by_css_selector("table[class='TableStandardBG']")
        driver.implicitly_wait(0)
        try:
            rows = table.find_element_by_css_selector("table[id='listingTable']>tbody").find_elements_by_css_selector("tr")
            for row in rows:
                if i != 0:
                    cells = row.find_elements_by_css_selector("td")
                    matchString = cells[8].text
                    while matchString[len(matchString)-1]==" ":
                        matchString = matchString[:len(matchString)-2]
                    m = re.compile(matchString[:4] + " ?0*" + matchString[5:], re.RegexFlag.IGNORECASE)
                    if cells[4].text=="Active":
                        for Xcontainer in containers:
                            if m.search(Xcontainer.number):
                                container=Xcontainer
                                break
                        j = 0
                        if container != "":
            #                 for cell in cells:
            #                     print(cell.text)
            #                     if j==1:
                            container.rvtime = cells[2].text + ' ' + cells[3].text
            #                     elif j==2:
            #                     elif j==5:
                            container.rvnumber = cells[6].text
                i+=1
        except:
            pass
        driver.implicitly_wait(600)
        driver.find_element_by_css_selector("img[src='/ImxEbusWeb/images/english/Back.gif']").click() 
        
        
def putInfoinExcel(containers, localContainers, switched):
    i = 0
    lastRun = switched["B1"].value
    currentRun = datetime.now()
    switched["B1"].value = currentRun
    lastRow = len(switched["A"])+1
    for row in localContainers.rows:
        if i != 0:
            if row[constants.CONTAINERCOL].value != None and row[constants.CONTAINERCOL].value != "":
                container = ""
                for Xcontainer in containers:
                    if Xcontainer.number==row[constants.CONTAINERCOL].value.upper():
                        container=Xcontainer
                        break
#                 print(row[0].fill.start_color.index)
                if row[0].fill.start_color.index==0:
                    pastMalport=False
                else:
                    pastMalport = row[0].fill.start_color.index[2:] == "FFFF00" or row[0].fill.start_color.index[2:] == "FF0000" 
#                 print(pastMalport)
                backupRow = []
                j=0
                for _ in row:
                    if row[j].value == "":
                        backupRow.append(None)
                    else:
                        backupRow.append(row[j].value)
                    j+=1
                row[constants.STATUSCOL].value=container.status
                if 'KC1' in container.size:
                    row[constants.SIZECOL].value="20DC"
                elif 'KC2' in container.size:
                    row[constants.SIZECOL].value="40DC"    
                elif 'KC4' in container.size:
                    row[constants.SIZECOL].value="40HC"
                elif 'KR1' in container.size:
                    row[constants.SIZECOL].value="20RF"
                elif 'KR4' in container.size:
                    row[constants.SIZECOL].value="40RH"
                row[constants.DOCUMENTSCOL].value=container.documents
                row[constants.CUSTOMSCOL].value=container.customs
                row[constants.STEAMSHIPCOL].value=container.steamship
                row[constants.OTHERCOL].value=container.other
                row[constants.ETACOL].value=container.eta
                row[constants.STORAGEDATECOL].value=container.storageDate
                row[constants.STORAGEAMOUNTCOL].value=container.storageAmount
                row[constants.RVTIMECOL].value=container.rvtime
                row[constants.RVNUMBERCOL].value=container.rvnumber
                
                malPort = False
                hold = False
                misc=False
                for i in range(len(row)):
                    
        #             if backupRow[i].value != None and backupRow[i].value != row[i].value:
#                     if i>len(row)-4:
#                         print(backupRow[i])
#                         print(row[i].value)
#                         print(backupRow[i]==row[i].value)
                    if i>9:
                        if backupRow[i] != row[i].value and not (row[i].value=="" and backupRow[i]==None):
                            row[i].fill = PatternFill(start_color='CCFFFF',
                            end_color='CCFFFF',
                            fill_type='solid')
                        else:
                            row[i].fill = PatternFill(start_color='FFFFFF',
                            end_color='FFFFFF',
                            fill_type='solid')
                        
                        if row[i].value and "Malport" in str(row[i].value):
                            malPort = True
                            
                        if row[i].value and "Mississauga" in str(row[i].value):
                            misc = True
                if row[constants.CUSTOMSCOL].value and "Hold" in row[constants.CUSTOMSCOL].value and ((row[constants.ETACOL].value and row[constants.ETACOL].value != " ") or (row[constants.STORAGEDATECOL].value and row[constants.STORAGEDATECOL].value != " ")):
                    hold =True
                
                fillColor = "FFFFFF"
                
                if malPort:
                    fillColor="FFFF00"
                if hold:
                    fillColor="FFC300"
                if misc:
                    fillColor="008000"
                if hold and (malPort or misc):
                    fillColor = "FF0000"
                
                
                for j in range(3):
                    row[j].fill = PatternFill(start_color=fillColor,
                    end_color=fillColor,
                    fill_type='solid')
                    
                if pastMalport != malPort:
                    blackFill=PatternFill(start_color='000000',
                            end_color='000000',
                            fill_type='solid')
                    switched.cell(lastRow+1, 1).value = lastRun
                    switched.cell(lastRow+2, 1).value = currentRun
                    switched.cell(lastRow, 1).fill = blackFill
                    for j in range(len(backupRow)):
                        switched.cell(lastRow, j+2).fill = blackFill
                        switched.cell(lastRow+1, j+2).value = backupRow[j]
                        switched.cell(lastRow+2, j+2).value = row[j].value
                    lastRow+=3
                    
        i+=1
        dims = {}
        for row in localContainers.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            localContainers.column_dimensions[col].width = value+3
#             switched.column_dimensions[col].width = value+3
        
        
def setupExcel(filepPath, containers):
      
    localContainersWb = load_workbook(filepPath, read_only=False, keep_vba=True)
    localContainers = localContainersWb["Current"]
    switched = localContainersWb["Switched"]
    for cell in range(1, localContainers.max_column):
#         print(cell)
        val = localContainers[1][cell].value
        if val=="Container": 
            constants.CONTAINERCOL=cell
        elif val=="Status": 
            constants.STATUSCOL=cell
        elif val=="Documents": 
            constants.DOCUMENTSCOL=cell
        elif val=="Customs": 
            constants.CUSTOMSCOL=cell
        elif val=="Steamship": 
            constants.STEAMSHIPCOL=cell
        elif val=="Other": 
            constants.OTHERCOL=cell
        elif val=="ETA": 
            constants.ETACOL=cell
        elif val=="1st Storage Date": 
            constants.STORAGEDATECOL=cell
        elif val=="Storage due": 
            constants.STORAGEAMOUNTCOL=cell
        elif val=="RV #": 
            constants.RVNUMBERCOL=cell
        elif val=="RV Time": 
            constants.RVTIMECOL=cell
        elif val=="Appointment Time": 
            constants.APPOINTMENTTIMECOL=cell
        elif val=="PB": 
            constants.PBCOL=cell
        elif val=="Size": 
            constants.SIZECOL=cell
    
    for cell in range(2, localContainers.max_row+1):
        contNumber = localContainers[cell][int(constants.CONTAINERCOL)].value
#         print(contNumber)
        if contNumber != "" and contNumber != None:
            container = Container(contNumber)
            container.checkDigit=localContainers[cell][int(constants.CONTAINERCOL)+1].value
            containers.append(container)
    
    return localContainersWb, localContainers, switched

# def copyRow(fromSheet, toSheet, indexCurrent, indexCompleted):
#     currentRow = fromSheet[indexCurrent]
#     completedRow = toSheet[indexCompleted]
#     for i in range(1, len(currentRow)):
#         new_cell=""
#         try:
#             new_cell = completedRow[i]
#         except:
#             new_cell = fromSheet.cell(row=indexCompleted, column=i)
#         cell=currentRow[i]
#         new_cell.value = currentRow[i].value
#         if cell.has_style:
#             new_cell.font = copy(cell.font)
#             new_cell.border = copy(cell.border)
#             new_cell.fill = copy(cell.fill)
#             new_cell.number_format = copy(cell.number_format)
#             new_cell.protection = copy(cell.protection)
#             new_cell.alignment = copy(cell.alignment)
    
# def moveCompleted(localContainersWb, filePath):
#     completed = localContainersWb["Completed"]
#     current = localContainersWb["Current"]
#     
# #     copyRow(current, completed, 1, 1)
#     destinationRow = completed.get_highest_row()
#     
#     for i in range(1, current.get_highest_row):
#         if "Load Out-Gate" in current.cell(row=i, column=7).value:
#             copyRow(current, completed, indexCurrent, indexCompleted) 
#     
#     
#     
#     dims = {}
#     for row in completed.rows:
#         for cell in row:
#             if cell.value:
#                 dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
#     for col, value in dims.items():
#         completed.column_dimensions[col].width = value+3
def checkContainerNumbers(containers):
    letterDictionary={"A":10,
                      "B":12,
                      "C":13,
                      "D":14,
                      "E":15,
                      "F":16,
                      "G":17,
                      "H":18,
                      "I":19,
                      "J":20,
                      "K":21,
                      "L":23,
                      "M":24,
                      "N":25,
                      "O":26,
                      "P":27,
                      "Q":28,
                      "R":29,
                      "S":30,
                      "T":31,
                      "U":32,
                      "V":34,
                      "W":35,
                      "X":36,
                      "Y":37,
                      "Z":38}
                      
    
    invalidContainerNumbers=""
    
    for container in containers:
        numberWithLeadingZeroes=container.number
        while len(numberWithLeadingZeroes)<10:
            numberWithLeadingZeroes = numberWithLeadingZeroes[:4]+"0"+numberWithLeadingZeroes[4:]
        try:
            sumCheck=letterDictionary[numberWithLeadingZeroes[0]]
            sumCheck+=letterDictionary[numberWithLeadingZeroes[1]]*2
            sumCheck+=letterDictionary[numberWithLeadingZeroes[2]]*4
            sumCheck+=letterDictionary[numberWithLeadingZeroes[3]]*8
            sumCheck+=int(numberWithLeadingZeroes[4])*16
            sumCheck+=int(numberWithLeadingZeroes[5])*32
            sumCheck+=int(numberWithLeadingZeroes[6])*64
            sumCheck+=int(numberWithLeadingZeroes[7])*128
            sumCheck+=int(numberWithLeadingZeroes[8])*256
            sumCheck+=int(numberWithLeadingZeroes[9])*512
            check = sumCheck%11
            if check==10:
                check=0
#         print(numberWithLeadingZeroes)
            if not check==int(container.checkDigit):
                invalidContainerNumbers+= numberWithLeadingZeroes + "\n"
        except:
            invalidContainerNumbers+= container.number + "\n"
    if invalidContainerNumbers!="":
#         popUpOK("Invalid container numbers (by check digit):\n"+invalidContainerNumbers)
#         sys.exit()
        
        
        top = Tk()
        top.config(bg = "lavender")
        L1 = Label(top, text="Invalid container numbers (by check digit):\n"+invalidContainerNumbers, bg="lavender", font=("serif", 16))
        L1.grid(row=0, column=0, columnspan=2)
        
        def callbackEnd():
            sys.exit()
             
        
        def callbackContinue():
            top.destroy()
             
        
        MyButton4 = Button(top, text="Proceed anyway", width=14, command=lambda: callbackContinue(), bg="green", font=("serif", 16))
        MyButton4.grid(row=2, column=0, padx=10, pady=10)
        
        MyButton5 = Button(top, text="Stop", width=25, command=lambda: callbackEnd(), bg="red", font=("serif", 16))
        MyButton5.grid(row=2, column=1, padx=10, pady=10)
        
        top.update()

        w = top.winfo_width() # width for the Tk root
        h = top.winfo_height() # height for the Tk root
           
        ws = top.winfo_screenwidth() # width of the screen
        hs = top.winfo_screenheight() # height of the screen
        x = (ws/2) - (w/2)
        y = (hs/2) - (h/2)
        
        top.geometry('%dx%d+%d+%d' % (w, h, x, y))
        
        top.update()

        top.lift()
        top.attributes('-topmost',True)
        top.after_idle(top.attributes,'-topmost',False)
#         moveTo(MyButton5.winfo_width()/2 + MyButton5.winfo_rootx(), MyButton5.winfo_height()/2 + MyButton5.winfo_rooty())
        
        top.mainloop()
            
if __name__ == '__main__':
#     try:
    filePath = r"J:\LOCAL DEPARTMENT\Automation - DO NOT MOVE\Incoming Local Containers.xlsm"
#     filePath = r"C:\Users\ssleep\Documents\Incoming Local Containers.xlsm"
#     filePath = r"C:\Users\Spencer\Documents\Incoming Local Containers.xlsm"
    
    containers = []
         
    localContainersWb, localContainers, switched = setupExcel(filePath, containers)

    checkContainerNumbers(containers)
    
    driver = setupCn()
    
    readContainerInfo(driver, containers)
    
    putInfoinExcel(containers, localContainers, switched)
#     except:
#         print(exc_info())
#         sleep(60)
    saved = False
    while not saved:
        try:
            localContainersWb.save(filePath)
            saved = True
        except:
            top = Tk()
            L1 = Label()
            L1 = Label(top, text="Please close \"Incoming Local Containers\" \nspreadsheet then hit \"OK\"")
            L1.config(font=("Courier", 16))
            L1.grid(row=0, column=0)
            top.lift()
            top.attributes('-topmost',True)
            top.after_idle(top.attributes,'-topmost',False)
              
            def callbackOK():
                top.destroy()
              
            MyButton5 = Button(top, text="OK", width=8, command=callbackOK)
            MyButton5.grid(row=1, column=0)
            MyButton5.config(font=("Courier", 16))
              
              
            w = 530 # width for the Tk root
            h = 100 # height for the Tk root
               
            # get screen width and height
            ws = top.winfo_screenwidth() # width of the screen
            hs = top.winfo_screenheight() # height of the screen
               
            # calculate x and y coordinates for the Tk root window
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
               
            # set the dimensions of the screen 
            # and where it is placed
            top.geometry('%dx%d+%d+%d' % (w, h, x, y))
           
            top.mainloop()
#     driver.quit()
#     driver.find_element_by_tag_name('html').send_keys(Keys.CONTROL, Keys.SHIFT, "w")
#     ActionChains(driver).send_keys(Keys.CONTROL, Keys.SHIFT, "w").perform()

    driver.quit()
#     sleep(1)
#     while True:
#         for p in process_iter():
#             if "geckodriver.exe" in p.name():
#                 p.kill()
#             break
    system('start excel.exe "'+ filePath + '"')
    
# pyinstaller "C:\Users\ssleep\workspace\CNLister\Lister\__init__.py" --distpath "J:\Spencer\CNLister" --noconsole -y

# pyinstaller "C:\Users\spencer\workspaceseaport\programs\CNLister\Lister\__init__.py" --distpath "c:\users\Spencer\CNLister" -y