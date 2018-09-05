from openpyxl import load_workbook
from sys import argv
from time import sleep


def moveCompleted(runningRouting, toMove):
    toMove = load_workbook(toMove)
    toMove = toMove.active
    
    containers=[]
    
    for row in toMove.rows:
        for cell in row:
            if cell.value and cell.value != "None":
                containers.append(formattedCont(cell.value))
    
    saveName = runningRouting
    runningRouting = load_workbook(runningRouting, read_only=False, keep_vba=True)
    incoming = runningRouting['Incoming']
    completed = runningRouting['Completed']
    
#     print(containers)
    
    toCutandPaste = []
#     i=0
    for row in incoming.rows:
#         print(row[2].value[:10])
        if len(row)>1 and row[2].value and row[2].value[:10] in containers:
            toCutandPaste.append(row)
    
    lastRow = len(completed["A"])
    
    j=1
    
    for row in toCutandPaste:
        for i in range(len(row)):
            completed.cell(row=lastRow+j, column=i+1).value = row[i].value
        j+=1
            
    for row in toCutandPaste:
        incoming.delete_rows(row[0].row)
    
    runningRouting.save(saveName)
    
def formattedCont(cont):
    if cont[4]== " ":
        cont = cont[:4] + cont[5:]
        
    while len(cont)<10:
        cont=cont[:4]+"0"+cont[5:]
    
    return cont[:10]
    

if __name__ == '__main__':
#     argv = r'["a",, ]'
    
#     folderPath = ''
#     for i in range(len(argv)):
#         if i!=0:
#             folderPath+=argv[i]
#             if i != len(argv) - 1:
#                 folderPath+=" "
#     folder1 = argv[1]
#     folder2 = argv[2]                
    
    folder1 = r"C:\Users\ssleep\Documents\Move Completed.xlsx"
    folder2 = r"C:\Users\ssleep\Documents\Running Routing 6-23a.xlsm"
    
    
    if "running" in folder1.lower():
        runningRouting = folder1
        toMove = folder2
    else:
        runningRouting = folder2
        toMove = folder1
        
    moveCompleted(runningRouting, toMove)
#     print(argv)
#     print(folder1)
#     print(folder2)
#     
#     sleep(100)
#     
    # pyinstaller "C:\Users\ssleep\workspace\MoveCompletedRunningRouting\Automator\__init__.py" --distpath "C:\Users\ssleep\Documents\movecompleted" -y