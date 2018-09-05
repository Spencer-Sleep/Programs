
from openpyxl import load_workbook

from sys import argv, exit

from pyautogui import moveTo, click, press, typewrite, hotkey

from win32api import GetKeyState

from tkinter import Button, Tk, Label, Entry, Frame, Radiobutton, StringVar,\
    constants

from time import sleep

import ContainerSizeInfo

import DispatchmateLocations as loc
from DispatchmateLocations import OKLOC, NEWLOC, SHIPPERLOC, DESCRIPTIONLOC,\
    CONSIGNEECODELOC, SHIPPERCODELOC, BOLLOC, OVERVIEWTABLOC, LANECODELOC
from HelperFunctions import done
import re
import HelperFunctions
from _datetime import datetime


contRates = {}
laneCodes = {}
laneCodesCoffee={}


CONTAINERNUMBER = "Container #"
PONUMBER = "WO"
# WEIGHLBS = "Weight (lbs)"
WEIGHT = "Weight"
PIECES = "Piece Count"
SIZE = "Size"
INBOND = "A8A?"
COFFEE= "COFFEE"
BOOKING = "BOL"
VESSEL = "Vessel"

def popUpOKLeft(text1, text2, textSize = 16):
    bgC = "lavender"
    top = Tk()
    top.config(bg = bgC)
    L1 = Label(top, text=text1, bg = bgC, padx = 20)
    L1.config(font=("serif", textSize))
    L1.grid(row=0, column=0, sticky=constants.W+constants.E)
    L1 = Label(top, text=text2, bg = bgC, padx = 20, justify=constants.LEFT)
    L1.config(font=("serif", textSize))
    L1.grid(row=1, column=0, sticky=constants.W + constants.E)
    def callbackOK():
#         sys.exit()
        top.destroy()
        
    MyButton = Button(top, text="OK", command=callbackOK)
    MyButton.grid(row=2, column=0, sticky=constants.W+constants.E, padx = 20, pady = (0,20))
    MyButton.config(font=("serif", 30), bg="green")
      
    top.update()
    
    w = top.winfo_width() # width for the Tk root
    h = top.winfo_height() # height for the Tk root
       
    ws = top.winfo_screenwidth() # width of the screen
    hs = top.winfo_screenheight() # height of the screen
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
    
    top.geometry('%dx%d+%d+%d' % (w, h, x, y))
    top.update()
    moveTo(MyButton.winfo_width()/2 + MyButton.winfo_rootx(), MyButton.winfo_height()/2 + MyButton.winfo_rooty())
    top.lift()
    top.attributes('-topmost',True)
    top.after_idle(top.attributes,'-topmost',False)
    top.mainloop()

def infoPopups():
#     steamShipLine = "HAMBURG"
    typeOfManifest = [-1]
    skip = [0]
    startAt = [""]
    top = Tk()
    L1 = Label(top, text="Please select which container to start at or \n how many containers to skip as well as\n whether to do PARS, A8As, or both \n ")
    L1.config(font=("Courier", 16))
    L1.grid(row=0, column=0, columnspan=3)
    L2 = Label(top, text="Start at container #:")
    L2.config(font=("Courier", 10))
    L2.grid(row=1, column=0)
    E1 = Entry(top, bd = 5, width = 30)
    E1.grid(row=1, column=1)
    L3 = Label(top, text="OR")
    L3.grid(row=2, column=1)
    L3.config(font=("Courier", 20))
    L4 = Label(top, text="# of containers to skip:")
    L4.grid(row = 3, column = 0)
    L4.config(font=("Courier", 10))
    E2 = Entry(top, bd = 5, width = 30)
    E2.grid(row=3, column=1)
    E2.insert(0, "0")
    
    
    steamShipLine=StringVar()
    
    f = Frame(top)
    f.grid(row=5, column=0, columnspan=3, pady=10)
    R1 = Radiobutton(f, text="HAM", variable=steamShipLine, value="HAMBURG", font=("Courier", 30), indicatoron=0, width = 7)
    R1.pack(side="left")
#     R1.config(font=("Courier", 16))
    R2 = Radiobutton(f, text="MSC", variable=steamShipLine, value="MSC", font=("Courier", 30), indicatoron=0, width = 7)
    R2.pack(side="left")
#     R2.config(font=("Courier", 16))
    R3 = Radiobutton(f, text="CMA", variable=steamShipLine, value="CMA", font=("Courier", 30), indicatoron=0, width = 7)
    R3.pack(side="left")
#     R3.config(font=("Courier", 16))
    R4 = Radiobutton(f, text="MAERSK", variable=steamShipLine, value="MAERSK", font=("Courier", 30), indicatoron=0, width = 7)
    R4.pack(side="left")
    
    steamShipLine.set("HAMBURG")
    
    top.lift()
    top.attributes('-topmost',True)
    top.after_idle(top.attributes,'-topmost',False)
      
    def callbackType(typeOf):
        startAt[0]=E1.get().strip()
        skip[0]=E2.get().strip()
        typeOfManifest[0] = typeOf
        top.destroy()
      
      
    
    MyButton5 = Button(top, text="PARS", command=lambda: callbackType(0), width=10)
    MyButton5.grid(row=4, column=0, pady=(10,0))
    MyButton5.config(font=("Courier", 16))
    MyButton6 = Button(top, text="A8A", command=lambda: callbackType(1), width=10)
    MyButton6.grid(row=4, column=1)
    MyButton6.config(font=("Courier", 16))
    MyButton7 = Button(top, text="BOTH", command=lambda: callbackType(2), width=10)
    MyButton7.grid(row=4, column=2)
    MyButton7.config(font=("Courier", 16))
      
      
    top.update()
    
    w = top.winfo_width() # width for the Tk root
    h = top.winfo_height() # height for the Tk root
       
    ws = top.winfo_screenwidth() # width of the screen
    hs = top.winfo_screenheight() # height of the screen
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
       
    # set the dimensions of the screen 
    # and where it is placed
    top.geometry('%dx%d+%d+%d' % (w, h, x, y))
   
    top.mainloop()
    
    steamShipLine = steamShipLine.get()
    
    terminal = [-1]
    top = Tk()
    L1 = Label(top, text="Please select which terminal these containers are coming from")
    L1.config(font=("Courier", 16))
    L1.grid(row=0, column=0, columnspan=5)
    top.lift()
    top.attributes('-topmost',True)
    top.after_idle(top.attributes,'-topmost',False)
      
    def callbackTerminal(terminalName):
#         startAt[0]=E1.get().strip()
#         skip[0]=E2.get().strip()
        terminal[0] = terminalName
        top.destroy()
    
    def callbackStop():
        top.destroy()
        exit()
    
    bgc = "lavender"
    
    MyButton5 = Button(top, text="APM", command=lambda: callbackTerminal("APM"), width=10, bg=bgc)
    MyButton5.grid(row=1, column=0)
    MyButton5.config(font=("Courier", 16))
    MyButton6 = Button(top, text="ASI", command=lambda: callbackTerminal("ASI"), width=10, bg=bgc)
    MyButton6.grid(row=1, column=1)
    MyButton6.config(font=("Courier", 16))
    MyButton7 = Button(top, text="GLOBAL", command=lambda: callbackTerminal("GLOBAL"), width=10, bg=bgc)
    MyButton7.grid(row=1, column=2)
    MyButton7.config(font=("Courier", 16))
    MyButton8 = Button(top, text="MAHER", command=lambda: callbackTerminal("MAHER"), width=10, bg=bgc)
    MyButton8.grid(row=1, column=3)
    MyButton8.config(font=("Courier", 16))
    MyButton8 = Button(top, text="NYCT", command=lambda: callbackTerminal("NYCT"), width=10, bg=bgc)
    MyButton8.grid(row=2, column=0)
    MyButton8.config(font=("Courier", 16))
    MyButton8 = Button(top, text="PACKER", command=lambda: callbackTerminal("PACKER"), width=10, bg=bgc)
    MyButton8.grid(row=2, column=1)
    MyButton8.config(font=("Courier", 16))
    MyButton8 = Button(top, text="PNCT", command=lambda: callbackTerminal("PNCT"), width=10, bg=bgc)
    MyButton8.grid(row=2, column=2)
    MyButton8.config(font=("Courier", 16))
    
    MyButton7 = Button(top, text="CSX", command=lambda: callbackTerminal("CSX"), width=10, bg="chocolate1")
    MyButton7.grid(row=3, column=0, pady=20)
    MyButton7.config(font=("Courier", 16))
    
    MyButton8 = Button(top, text="STOP", command=lambda: callbackStop(), width=20, bg="red")
    MyButton8.grid(row=4, column=0, pady=10, columnspan=4)
    MyButton8.config(font=("Courier", 24))
      
    top.update()
    
    w = top.winfo_width() # width for the Tk root
    h = top.winfo_height() # height for the Tk root
       
    ws = top.winfo_screenwidth() # width of the screen
    hs = top.winfo_screenheight() # height of the screen
    
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
       
    # set the dimensions of the screen 
    # and where it is placed
    top.geometry('%dx%d+%d+%d' % (w, h, x, y))
   
    top.mainloop() 
    
    return startAt, skip, terminal[0], steamShipLine, typeOfManifest

def loadContainerInfo(specificPath, startAt, skip, terminal, steamShipLine):
    global CONTAINERNUMBER
    global PONUMBER 
    global WEIGHT
    global PIECES
    global SIZE
    global INBOND
    global COFFEE
    global BOOKING
    global VESSEL
    
    
    
    routingBook = load_workbook(specificPath)
    routing = routingBook.active
    
    csxFormat = True
    for cell in next(routing.rows):
        if cell.value and "A8A?" in cell.value.upper():
            csxFormat=False
        
        
    pars = []
    a8a = []
    if csxFormat:
        CONTAINERNUMBER = "Container"
#         PONUMBER = "770505998"
        # WEIGHLBS = "Weight (lbs)"
#         WEIGHT = "Weight"
        PIECES = "Piece Count"
        INBOND = "Border"
        COFFEE= "Commodity"
        BOOKING = "TPDoc"
        PONUMBER = "Shipper"
    
    colDict = {CONTAINERNUMBER: "",
           PONUMBER: "",
#                WEIGHLBS: "",
           WEIGHT: "",
           SIZE: "",
           INBOND: "",
           COFFEE: "",
           BOOKING: "",
           VESSEL: "",
           PIECES:""}   
    
    payoutCol = ""
    terminalCol = ""
    SSLCol = ""
    
    
    for cell in next(routing.rows):
        if cell.value:
            for contProperty in colDict:
                if contProperty in cell.value:
                    colDict[contProperty] = cell.col_idx - 1
            if "PAYOUT TIER" in cell.value:
                payoutCol = cell.col_idx - 1
            if "Terminal" in cell.value:
                terminalCol = cell.col_idx - 1
            if "SSL" in cell.value:
                SSLCol = cell.col_idx - 1
        lastIndex = cell.col_idx
    if csxFormat and not terminal=="CSX":
        payoutCol = lastIndex+1
    
    popUpMessage = ""
    for column, content in [(CONTAINERNUMBER, "container number"), (SIZE, "size"), (INBOND, "clearance type (A8A vs PARS)"), (BOOKING, "booking"), (VESSEL, "vessel"), (PONUMBER, "PO/WO number"), (WEIGHT, "weight")]:
        if colDict[column]=="":
            popUpMessage = popUpMessage + "Could not find a column named: \"" + column + "\",\n which should contain the " + content + ".\n\n"
    if popUpMessage != "":
        HelperFunctions.popUpOK(popUpMessage)
        exit()
    i=0
    found = startAt[0]==""
    containerList = [""]
    
    class Container(object):
        def __init__(self):
            self.properties = {CONTAINERNUMBER: "",
                   PONUMBER: "",
    #                WEIGHLBS: "",    
                   WEIGHT: "",
                   PIECES: "",
                   SIZE: "",
                   INBOND: "",
                   COFFEE: "",
                   BOOKING: "",
                   VESSEL: ""}
            
    hiddenRows = []
    for rowNum, rowDimension in routing.row_dimensions.items():
        if rowDimension.hidden == True:
            hiddenRows.append(rowNum)   
    pattern = re.compile(r"[A-Z]{4}[0-9]{7}")
    
    invalids = "The following values are invalid:"
    invalidConts = ""
    invalidWeight=""
    invalidSizes=""
    invalidPieces=""
    
    try:
        for row in routing.rows:
            if i>int(skip[0]) and not (i+1) in hiddenRows:
                container = Container()
                for contProperty in container.properties:
    #                 if contProperty != WEIGHLBS:    
    #                 if colDict[contProperty] != "":
                    if colDict[contProperty] != "": 
                        container.properties[contProperty] = str(row[colDict[contProperty]].value).upper()
    #                 if contProperty == INBOND:
    #                     print(colDict[contProperty])
    #                     print(str(row[colDict[contProperty]].value))
                if found or container.properties[CONTAINERNUMBER]==startAt[0]:
#                     m = re.search(r'[A-Za-z]{4}\d{7}', container.properties[CONTAINERNUMBER])
                    notBlank = False
                    for prop in container.properties.values():
                        notBlank = (prop != "NONE" and prop != "") or notBlank
                    if not container.properties[CONTAINERNUMBER] in containerList and notBlank:
                        found = True

                        containerList.append(container.properties[CONTAINERNUMBER])
                        if not pattern.fullmatch(container.properties[CONTAINERNUMBER]):
                            invalidConts = invalidConts + ("\nContainer number: " + container.properties[CONTAINERNUMBER])
                        if container.properties[PIECES] and container.properties[PIECES]!= "NONE":
                            container.properties[PIECES] = container.properties[PIECES].split(" ")[0]
                            try:
                                int(float(container.properties[PIECES]))
                            except:
                                invalidPieces = invalidPieces + "\nPiece count for container " + container.properties[CONTAINERNUMBER]+ ": " + container.properties[PIECES]
                        
                        if steamShipLine=="MAERSK":
                            if not ("Coffee, non-roasted, non-frozen" in container.properties[COFFEE]):
                                container.properties[COFFEE]="NONE"
                        elif csxFormat:
                            if not "coffee" == container.properties[COFFEE].strip().lower():
                                container.properties[COFFEE]="NONE"
                        if container.properties[WEIGHT] != "NONE":
                            try:
                                weight = int(float(container.properties[WEIGHT]))
                            except:
                                invalidWeight = invalidWeight + "\nWeight for container " + container.properties[CONTAINERNUMBER]+ ": " +container.properties[WEIGHT]
                            
                        size = ContainerSizeInfo.standardSize(container.properties[SIZE])
                        if size:
                            skipThisContainer = [False]
                            
                            if weight > contRates[size].CSXlimit:
                                top = Tk()
                                top.config(bg = "lavender")
                                L1 = Label(top, text="Container " + container.properties[CONTAINERNUMBER] + " is too heavy for Buffalo.\nThe weight limit for "+ size + 
                                           " is " + str(contRates[size].CSXlimit) + " KG.\n" + container.properties[CONTAINERNUMBER] + " is " + str(weight) + " KG.", bg="lavender", font=("serif", 16))
                                L1.grid(row=0, column=0, columnspan=2)
                                
                                def callbackSkip(skipThisContainer):
                                    skipThisContainer[0] = True
                                    top.destroy() 
                                
                                def callbackContinue():
                                    top.destroy()
                                     
                                
                                MyButton4 = Button(top, text="Proceed anyway", width=14, command=lambda: callbackContinue(), bg="green", font=("serif", 16))
                                MyButton4.grid(row=2, column=0, padx=10, pady=10)
                                
                                MyButton5 = Button(top, text="Skip this container", width=25, command=lambda: callbackSkip(skipThisContainer), bg="red", font=("serif", 16))
                                MyButton5.grid(row=2, column=1, padx=10, pady=10)
                                
                                top.update()
        
                                w = top.winfo_width() # width for the Tk root
                                h = top.winfo_height() # height for the Tk root
                                   
                                ws = top.winfo_screenwidth() # width of the screen
                                hs = top.winfo_screenheight() # height of the screen
                                x = (ws/2) - (w/2)
                                y = (hs/2) - (h/2)
                                
                                top.geometry('%dx%d+%d+%d' % (w, h, x, y))
                                
                                top.update()
        
                                top.lift()
                                top.attributes('-topmost',True)
                                top.after_idle(top.attributes,'-topmost',False)
                                moveTo(MyButton5.winfo_width()/2 + MyButton5.winfo_rootx(), MyButton5.winfo_height()/2 + MyButton5.winfo_rooty())
                                
                                top.mainloop()
                            
                            if not skipThisContainer[0]:
                                if container.properties[INBOND].lower() != "pars" and container.properties[INBOND]!="NONE":
                                    a8a.append(container)
                                else:
                                    pars.append(container)
                            
                            if terminal!="CSX": 
                                print(terminal)
                                tier = ""
                                if container.properties[COFFEE] != "NONE": 
                                    tier = "TIER 1"
                                else:
                                    if weight>contRates[size].T4weight:
                                        tier = "THRUWAY"
                                    elif weight>contRates[size].T3weight:
                                        tier = "TIER 3"
                                    elif weight>contRates[size].T2weight:
                                        tier = "TIER 2"
                                    elif weight>contRates[size].T1weight:
                                        tier = "TIER 1"
                                    else:
                                        tier = ""
                                if csxFormat:
                                    routing.cell(row=row[0].row, column=payoutCol).value = tier
                                else:
                                    row[payoutCol].value = tier
                            if not csxFormat:
                                row[terminalCol].value = terminal
                                row[SSLCol].value = steamShipLine
                    
                        else:
                            invalidSizes = invalidSizes + "\nSize/type for container " + container.properties[CONTAINERNUMBER]+ ": " +container.properties[SIZE]
                        
#                         ccnPattern.fullmatch(container.properties[CCN])
            i+=1
        invalidExtras = invalidConts + invalidWeight+invalidSizes+invalidPieces
        if invalidExtras != "":
            popUpOKLeft(invalids, invalidExtras)
            raise AssertionError
    except AssertionError:
        exit()
    except:
        try:
            top.destroy()
        except:
            pass
        HelperFunctions.popUpOK("Something went wrong reading data from the Spreadsheet.\n Make sure everything has a valid weight and size\n and container number, then start again.", 16)
        raise
        exit()
    if not terminal=="CSX":
        try:
            routingBook.save(specificPath)
        except:  
            routingBook.save(specificPath[:-5] + " With Info and Tiers" + specificPath[-5:])
        
    return pars, a8a      
    
def bookPars(containers, terminal, steamShipLine):
    terminalNum = ""
    if terminal == 'NYCT':
        terminalNum = '664'
    elif terminal == 'GLOBAL':
        terminalNum = '304'
    elif terminal == 'PACKER':
        terminalNum = '309'
    elif terminal == 'APM':
        terminalNum = '306'
    elif terminal == 'ASI':
        terminalNum = '305'
    elif terminal == 'PNCT':
        terminalNum = '310'
    elif terminal == 'MAHER':
        terminalNum = '330'
    elif terminal == 'CSX':
        terminalNum = '311'
    
    click(NEWLOC)
    press("tab")
    press("enter")
    
    click(loc.SHIPPERCODELOC)
    typewrite(terminalNum)
    press('tab')
     
    click(loc.CONSIGNEECODELOC)
    typewrite("303")
    press('tab')
     
    click(loc.CUSTOMERCODELOC)
    if steamShipLine=="HAMBURG":
        typewrite("117")
    elif steamShipLine=="MSC":
        if terminal=="CSX":
            typewrite("1635")
        else:
            typewrite("520")
    elif steamShipLine=="CMA":
        typewrite("1731")
    elif steamShipLine=="MAERSK":
        if terminal=="CSX":
            typewrite("1779")
        else:
            typewrite("411")
    press('tab')
    
    click(loc.DESCRIPTIONLOC)
    if steamShipLine=="MSC" or (steamShipLine=="HAMBURG" and terminal=="CSX"):
        typewrite("IMPORT")
    press("tab", 8)
    typewrite("r")
    if((steamShipLine =="MSC" or steamShipLine=="HAMBURG") and terminal=="CSX"):
        typewrite("t")
    
    press("tab", 4)
    typewrite(steamShipLine)
    
    click(loc.DIRECTIONLOC)
    typewrite("imp")
    press("enter")
    
    click(loc.DIVISIONLOC)
    if terminal != "CSX":
        typewrite("usa")
    else:
        typewrite("TOR")
    press("enter")
    
    
    click(loc.HOUSELOC)
    hotkey('ctrl', 'a')
    if terminal=="CSX":
        typewrite("REGIONAL")
    else:
        typewrite("LONG")
    
    lastContainerRates = 0
    
    for container in containers:
        if terminal=="PACKER" and contRates[ContainerSizeInfo.standardSize(container.properties[SIZE])].P2weight and int(float(container.properties[WEIGHT]))>contRates[ContainerSizeInfo.standardSize(container.properties[SIZE])].P2weight:
            top = Tk()
            L1 = Label(top, text="Container " + container.properties[CONTAINERNUMBER] + " is too heavy for PA \n Skipping container.")
            L1.grid(row=0, column=0)
             
            def callbackTerminal(terminalNum):
                top.destroy()
             
            MyButton4 = Button(top, text="OK", width=10, command=lambda: callbackTerminal(terminalNum))
            MyButton4.grid(row=1, column=0)
             
            top.lift()
            top.attributes('-topmost',True)
            top.after_idle(top.attributes,'-topmost',False)
             
            top.update()
    
            w = top.winfo_width() # width for the Tk root
            h = top.winfo_height() # height for the Tk root
               
            ws = top.winfo_screenwidth() # width of the screen
            hs = top.winfo_screenheight() # height of the screen
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
            
            top.geometry('%dx%d+%d+%d' % (w, h, x, y))
             
            # set the dimensions of the screen 
            # and where it is placed
             
            moveTo(968, 561)
#             def click2():
#                 click(1001, 508)
#              
#             top.after(10, click2)
            top.mainloop()
        else:
            click(loc.OVERVIEWTABLOC)
            
            click(loc.OKLOC)
            click(loc.DUPLICATELOC)
            
            if GetKeyState(27) < 0:
                exit()
            
            if not (steamShipLine=="HAMBURG" and terminal!="CSX"):
                click(loc.BOLLOC)
                if steamShipLine=="MAERSK":
                    typewrite(container.properties[BOOKING])
                else:
                    typewrite(container.properties[VESSEL])
            
            if not steamShipLine=="MSC" and not steamShipLine=="CMA":
                click(loc.CONSIGNEETRACELOC)
                if steamShipLine=="HAMBURG":
                    typewrite(container.properties[PONUMBER])
                elif steamShipLine=="MAERSK":
                    typewrite(container.properties[VESSEL])
                else:
                    typewrite(container.properties[BOOKING])
            
            if not (steamShipLine=="MAERSK"):
                click(loc.POLOC)
                if (steamShipLine=="MSC" and terminal!= "CSX"):
                    typewrite(container.properties[BOOKING])
                else:
                    typewrite(container.properties[PONUMBER])
            
            click(loc.DESCRIPTIONLOC)
            press("tab", 7)
            typewrite("1")
            
            press("tab", 2)
            typewrite(container.properties[CONTAINERNUMBER])
            
            
            click(loc.EQUIPMENTLOC)
            press('home')
            size = ContainerSizeInfo.standardSize(container.properties[SIZE])
            if size == "20D86":
                downAmount = 4
            elif size == "20R86":
                downAmount = 3
            elif size == "20O86":
                downAmount = 2
            elif size == "40D86":
                downAmount = 7
            elif size == "40O86":
                downAmount = 11
            elif size == "40D96":
                downAmount = 10
            elif size == "40R96":
                downAmount = 12
            elif size == "D96":
                downAmount = 13
            
            press('down', downAmount)
            press('enter')
            
            click(loc.LANECODELOC)
            
            if container.properties[COFFEE] != "NONE" and laneCodesCoffee[terminal]==556:
                typewrite("56")
                press('up')
            elif container.properties[COFFEE] != "NONE":
                typewrite(str(laneCodesCoffee[terminal]))
            elif terminal !="CSX" and int(float(container.properties[WEIGHT]))>contRates[size].T4weight and laneCodesThru[terminal]==555:
                typewrite("56")
                press("up", 2)
            elif terminal !="CSX" and int(float(container.properties[WEIGHT]))>contRates[size].T4weight:
                typewrite(laneCodesThru[terminal])
            elif laneCodes[terminal]==554:
                typewrite("56")
                press('up', 3)
            else:
                typewrite(str(laneCodes[terminal]))
            press("enter")
    
            
    #   *****************************
    #        Routing Tab
    #   *****************************
    
            if terminal != "CSX":
                overweight = False
                reefer = False
                click(loc.ROUTINGTABLOC)
                
                payout = 0
                OW = False
                
                weight = int(float(container.properties[WEIGHT]))
                
                
                if weight>contRates[size].P1weight:
                    overweight = True
                if "R" in size:
                    reefer = True
                
                
                if container.properties[COFFEE] != "NONE": 
                    payout = contRates[size].T1payout
                else:
                    if weight>contRates[size].T4weight:
                        payout = contRates[size].T4payout
                        OW = True
                    elif weight>contRates[size].T3weight:
                        payout = contRates[size].T3payout
                    elif weight>contRates[size].T2weight:
                        payout = contRates[size].T2payout
                    elif weight>contRates[size].T1weight:
                        payout = contRates[size].T1payout
                
                if payout != 0:
                    click(loc.DRIVERPAYOUTLOC)
                    typewrite("over")
                    
                    press("tab", 3)
                      
        #             f=open("J:\Spencer\Hapag-Lloyd Dispatchmate\Overweights.txt")
        #             overweights = []
        #             for i in range(6):
        #                 overweights.append(f.readline())
                    
                    typewrite(str(payout))
                    if OW:
                        click(loc.DRIVERPAYOUTLOC[0], loc.DRIVERPAYOUTLOC[1]+19)
                        typewrite("thru")
                        press("tab", 3)
                        typewrite(str(contRates[size].OWpayout))
            
            
            
    #   *****************************
    #        Rating Tab
    #   *****************************
       
                click(loc.RATINGTABLOC)
                 
                clickHeight = 0
                 
        #         click(89, clickHeight)
        #         clickHeight += 20
        #         typewrite("ont")
        #         press("tab")
        #         if terminalNum[0]=="309":
        #             typewrite("946")
        #         else:
        #             typewrite("858")
                
                rateCount = 0
                
                if(reefer):
                    click(loc.CUSTOMERCHARGELOC[0], loc.CUSTOMERCHARGELOC[1]+ clickHeight)
                    clickHeight += 19
                    hotkey('ctrl', 'a')
                    typewrite('re')
                    press("down")
        #             press("tab", 3)
        #             hotkey('ctrl', 'a')
        #             typewrite(contRates[size].reeferCharge)
                    rateCount+=1
                 
                if(overweight and container.properties[COFFEE]=="NONE"):
                    click(loc.CUSTOMERCHARGELOC[0], loc.CUSTOMERCHARGELOC[1]+ clickHeight)
                    clickHeight += 19
                    hotkey('ctrl', 'a')
                    typewrite("ov")
                    press("tab")
                    if contRates[size].P2weight and weight>contRates[size].P2weight:
                        press("tab", 2)
                        hotkey('ctrl', 'a')
                        typewrite(str(contRates[size].P2rate))
                    
                    rateCount+=1
                    
                if terminalNum=="664":
                    click(loc.CUSTOMERCHARGELOC[0], loc.CUSTOMERCHARGELOC[1]+ clickHeight)
                    clickHeight += 19
                    hotkey('ctrl', 'a')
                    typewrite("n")
                    press("tab")
                    rateCount+=1
                
                if rateCount<lastContainerRates:
                    for _ in range(lastContainerRates-rateCount):
                        click(loc.CUSTOMERCHARGELOC[0], loc.CUSTOMERCHARGELOC[1]+clickHeight)
                        clickHeight += 19
                        hotkey('ctrl', 'a')
                        press("del")
                        press("tab")
                        press("del")
        
                lastContainerRates = rateCount
            click(loc.OKLOC)
        
def bookA8A(containers, terminal, steamShipLine):
    terminalNum = ""
    if terminal == 'NYCT':
        terminalNum = '664'
    elif terminal == 'GLOBAL':
        terminalNum = '304'
    elif terminal == 'PACKER':
        terminalNum = '309'
    elif terminal == 'APM':
        terminalNum = '306'
    elif terminal == 'ASI':
        terminalNum = '305'
    elif terminal == 'PNCT':
        terminalNum = '310'
    elif terminal == 'MAHER':
        terminalNum = '330'
    elif terminal == 'CSX':
        terminalNum = '311'
    
#     if typeOfManifest[0]==1:
    click(NEWLOC)
    press("tab")
    press("enter")
    
    
    click(loc.CUSTOMERCODELOC)
    if steamShipLine=="HAMBURG":
        typewrite("1788")
    elif steamShipLine=="MSC":
        if terminal=="CSX":
            typewrite("1635")
        else:
            typewrite("520")
    elif steamShipLine=="CMA":
        typewrite("1731")
    elif steamShipLine=="MAERSK":
        if terminal=="CSX":
            typewrite("1779")
        else:
            typewrite("411")
    press('tab')

    click(SHIPPERCODELOC)
    typewrite(terminalNum)
    press("tab")

    click(CONSIGNEECODELOC)
    typewrite("303")
    press("tab")

#     press("tab", 4)
#     typewrite("MAERSK")

    click(loc.DIRECTIONLOC)
    typewrite("imp")
    press("enter")
    
    click(loc.DIVISIONLOC)
    if terminal != "CSX":
        typewrite("usa")
    else:
        typewrite("TOR")
    press("enter")                      


#         click(loc.LANECODELOC)
#         if laneCodes[terminal]==554:
#             typewrite("56")
#             press('up', 3)
#         else:
#             typewrite(str(laneCodes[terminal]))
#         press("enter")
#     
    click(loc.HOUSELOC)
    hotkey('ctrl', 'a')
    if terminal=="CSX":
        typewrite("REGIONAL")
    else:
        typewrite("LONG")
#     elif steamShipLine=="HAMBURG":
#         click(OVERVIEWTABLOC)
#         click(loc.CUSTOMERCODELOC)
#         typewrite("1788")
#         press("tab")
#     click(1153, 98)
#     click(89, 144)
#     typewrite("on")
#     press("tab")
#     
#     click(248, 101)
    
    click(loc.DOCUMENTATIONANDCUSTOMSTABLOC)
    
    click(loc.LOCATIONOFGOODSLOC)
    typewrite("4")
    if terminal != "CSX":
        typewrite("2")
    press("tab")
    typewrite("0")
    press("tab")
    typewrite("0")
    
    click(loc.CROSSINGLOC)
    hotkey('ctrl', 'a')
    typewrite("0901")
    press("enter")
    
    click(loc.CONTROLLOC)
    typewrite("20C0")
    
    if typeOfManifest[0]==1:
        lastContainerRates=4
    else:
        lastContainerRates=1
    for container in containers:
        if terminal=="PACKER" and contRates[ContainerSizeInfo.standardSize(container.properties[SIZE])].P2weight and int(float(container.properties[WEIGHT]))>contRates[ContainerSizeInfo.standardSize(container.properties[SIZE])].P2weight:
            top = Tk()
            L1 = Label(top, text="Container " + container.properties[CONTAINERNUMBER] + " is too heavy for PA \n Skipping container.")
            L1.grid(row=0, column=0)
             
            def callbackTerminal(terminalNum):
                top.destroy()
             
            MyButton4 = Button(top, text="OK", width=10, command=lambda: callbackTerminal(terminalNum))
            MyButton4.grid(row=1, column=0)
             
            top.lift()
            top.attributes('-topmost',True)
            top.after_idle(top.attributes,'-topmost',False)
             
            w = 300 # width for the Tk root
            h = 150 # height for the Tk root
             
            # get screen width and height
            ws = top.winfo_screenwidth() # width of the screen
            hs = top.winfo_screenheight() # height of the screen
             
            # calculate x and y coordinates for the Tk root window
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
             
            # set the dimensions of the screen 
            # and where it is placed
            top.geometry('%dx%d+%d+%d' % (w, h, x, y))
             
            moveTo(1001, 508)
#             def click2():
#                 click(1001, 508)
#              
#             top.after(10, click2)
            top.mainloop()
        else:
            click(loc.OVERVIEWTABLOC)
            
            click(loc.OKLOC)
            click(loc.DUPLICATELOC)
            
            click(SHIPPERLOC)
            
            if GetKeyState(27) < 0:
                exit()
            
            while not GetKeyState(45)<0:
    #             if not GetKeyState(71)<0:
                True
            
            if GetKeyState(27) < 0:
                exit()
                
            if not steamShipLine=="HAMBURG":
                click(loc.BOLLOC)
                if steamShipLine=="MAERSK":
                    typewrite(container.properties[BOOKING])
                else:
                    typewrite(container.properties[VESSEL])
            
            if not steamShipLine=="MSC" and not steamShipLine=="CMA":
                click(loc.CONSIGNEETRACELOC)
                if steamShipLine=="HAMBURG":
                    typewrite(container.properties[PONUMBER])
                elif steamShipLine=="MAERSK":
                    typewrite(container.properties[VESSEL])
                else:
                    typewrite(container.properties[BOOKING])
            
            if not (steamShipLine=="MAERSK"):
                click(loc.POLOC)
                if (steamShipLine=="MSC" and terminal!= "CSX"):
                    typewrite(container.properties[BOOKING])
                else:
                    typewrite(container.properties[PONUMBER])
            
#             click(loc.POLOC)
#             typewrite(container.properties[PONUMBER])
            
    #         click(60, 134)
    #         press("tab")
            
            click(loc.DESCRIPTIONLOC)
            press("tab", 1)
            typewrite(container.properties[WEIGHT])
            
            press("tab", 6)
            if container.properties[PIECES]!="NONE" and container.properties[PIECES]!="":
                typewrite(container.properties[PIECES])
            
            press("tab")
            typewrite("r")
            if(steamShipLine =="MSC" and terminal=="CSX"):
                typewrite("t")
    #         press("tab", 6)
            
    #         click(550, 350)
    #         typewrite(container.properties[PIECES])
            
    #         click(18, 351)
            
            
            press("tab")
            typewrite(container.properties[CONTAINERNUMBER])
            
            size = ContainerSizeInfo.standardSize(container.properties[SIZE])
    #         if size == "20DC or 20st":
    #             sizeCode = "20DC"
    #         elif size == "20RF":
    #             sizeCode = "20RF"
    #         elif size == "40 DRY 8 6":
    #             sizeCode = "40DC"
    #         elif size == "40 DRY 9 6":
    #             sizeCode = "40HC"
    #         elif size == "40 REEF 9 6":
    #             sizeCode = "40RH"
    #         elif size == " DRY 9 6":
    #             sizeCode = "DC"
                
                
            press("tab", 3)
            typewrite(steamShipLine + " " + size)
            
    #         click(loc.BOLLOC)
    #         typewrite(container.properties[BOOKING])
            
#             click(472, 304)
#             typewrite(container.properties[VESSEL])
            
    #         click(18, 351)
    #         press("tab", 7)
    #         typewrite("1")
            
            
            
            
            click(loc.EQUIPMENTLOC)
            press('home')
            
            downAmount = 0
            
            if size == "20D86":
                downAmount = 4
            elif size == "20R86":
                downAmount = 3
            elif size == "20O86":
                downAmount = 2
            elif size == "40D86":
                downAmount = 7
            elif size == "40O86":
                downAmount = 11
            elif size == "40D96":
                downAmount = 10
            elif size == "40R96":
                downAmount = 12
            elif size == "D96":
                downAmount = 13
            
            
            
            press('down', downAmount)
            press('enter')
            
            click(LANECODELOC)
            if container.properties[COFFEE] != "NONE" and laneCodesCoffee[terminal]==556:
                typewrite("56")
                press('up')
            elif container.properties[COFFEE] != "NONE":
                typewrite(str(laneCodesCoffee[terminal]))
            elif terminal !="CSX" and int(float(container.properties[WEIGHT]))>contRates[size].T4weight and laneCodesThru[terminal]==555:
                typewrite("56")
                press("up", 2)
            elif terminal !="CSX" and int(float(container.properties[WEIGHT]))>contRates[size].T4weight:
                typewrite(laneCodesThru[terminal])
            elif laneCodes[terminal]==554:
                typewrite("56")
                press('up', 3)
            else:
                typewrite(str(laneCodes[terminal]))
            press("enter")
            
            
            
    #   *****************************
    #        Routing Tab
    #   *****************************
    
            if terminal != "CSX":
                overweight = False
                reefer = False
                click(loc.ROUTINGTABLOC)
                
                payout = 0
                OW = False
                
                weight = int(float(container.properties[WEIGHT]))
                
                
                if weight>contRates[size].P1weight:
                    overweight = True
                if "R" in size:
                    reefer = True
                
                
                if container.properties[COFFEE] != "NONE": 
                    payout = contRates[size].T1payout
                else:
                    if weight>contRates[size].T4weight:
                        payout = contRates[size].T4payout
                        OW = True
                    elif weight>contRates[size].T3weight:
                        payout = contRates[size].T3payout
                    elif weight>contRates[size].T2weight:
                        payout = contRates[size].T2payout
                    elif weight>contRates[size].T1weight:
                        payout = contRates[size].T1payout
                
                if payout != 0:
                    click(loc.DRIVERPAYOUTLOC)
                    typewrite("over")
                    
                    press("tab", 3)
                      
        #             f=open("J:\Spencer\Hapag-Lloyd Dispatchmate\Overweights.txt")
        #             overweights = []
        #             for i in range(6):
        #                 overweights.append(f.readline())
                    
                    typewrite(str(payout))
                    if OW:
                        click(loc.DRIVERPAYOUTLOC[0], loc.DRIVERPAYOUTLOC[1]+19)
                        typewrite("thru")
                        press("tab", 3)
                        typewrite(str(contRates[size].OWpayout))
            
            
            
    #   *****************************
    #        Rating Tab
    #   *****************************
     
            
                click(loc.RATINGTABLOC)
                 
                clickHeight = 0
                
                rateCount = 0
                
                if(reefer):
                    click(loc.CUSTOMERCHARGELOC[0], loc.CUSTOMERCHARGELOC[1]+ clickHeight)
                    clickHeight += 19
                    hotkey('ctrl', 'a')
                    typewrite('re')
                    press("down")
    #                 press("tab", 3)
    #                 hotkey('ctrl', 'a')
    #                 typewrite(str(contRates[size].reeferCharge))
        #             print(contRates[size].reeferCharge)
                    rateCount+=1
                
                if(overweight and container.properties[COFFEE]=="NONE"):
                    click(loc.CUSTOMERCHARGELOC[0], loc.CUSTOMERCHARGELOC[1]+ clickHeight)
                    clickHeight += 19
                    hotkey('ctrl', 'a')
                    typewrite("ov")
                    press("tab")
                    if contRates[size].P2weight and weight>contRates[size].P2weight:
                        press("tab", 2)
                        hotkey('ctrl', 'a')
                        typewrite(str(contRates[size].P2rate))
                    
                    rateCount+=1
                    
                if terminalNum[0]=="664":
                    click(loc.CUSTOMERCHARGELOC[0], loc.CUSTOMERCHARGELOC[1]+ clickHeight)
                    clickHeight += 19
                    hotkey('ctrl', 'a')
                    typewrite("n")
                    press("tab")
                    rateCount+=1
                
                if rateCount<lastContainerRates:
                    for _ in range(lastContainerRates-rateCount):
                        click(loc.CUSTOMERCHARGELOC[0], loc.CUSTOMERCHARGELOC[1]+clickHeight)
                        clickHeight += 19
                        hotkey('ctrl', 'a')
                        press("del")
                        press("tab")
                        press("del")
        
                lastContainerRates = rateCount
            
            click(loc.OKLOC)


if __name__ == '__main__':
#     argv = r"a J:\Running Routing by Vessel\GERDA MAERSK 830W.xlsx".split()
#     argv = r"a J:\All motor routings\2018\Week 29\HAMBURG\MONTE ACONCAGUA V-82N\PA\PBs.xlsx".split()

#     print("NONE".split(" ")[0])
#     print("1234 asd".split(" ")[0])
#     print("1234asd".split(" ")[0])
#     exit()
    specificPath = ''
    for i in range(len(argv)):
        if i!=0:
            specificPath+=argv[i]
            if i != len(argv) - 1:
                specificPath+=" "
    
    startAt, skip, terminal, steamShipLine, typeOfManifest = infoPopups()
    P1str, P2str = "",""
    if steamShipLine=="HAMBURG":
        P1str, P2str = "HAMBURG SUD OW 1", "HAMBURG SUD OW 2"
    elif steamShipLine=="MSC":
        P1str, P2str = "MSC", ""
    elif steamShipLine=="CMA":
        P1str, P2str = "CMA", "",
    elif steamShipLine=="MAERSK":
        P1str, P2str = "MAERSK", "",
        
    values = ContainerSizeInfo.loadValues(True, True, P1str, P2str, True)
    contRates = values[0]
    laneCodes = values[1]
    laneCodesCoffee = values[2]
    laneCodesThru = values[3]
    
    pars, A8A = loadContainerInfo(specificPath, startAt, skip, terminal, steamShipLine)
    
    if (typeOfManifest[0] == 0 or typeOfManifest[0] == 2) and len(pars)>0:
        bookPars(pars, terminal, steamShipLine)
    if typeOfManifest[0] > 0 and len(A8A)>0:
        bookA8A(A8A, terminal, steamShipLine)
    
    done()
    
# pyinstaller "C:\Users\ssleep\workspace\Hamburg Sud Dispatchmate\Automator\__init__.py" --distpath "J:\Spencer\Hamburg Sud Dispatchmate" --noconsole -y