from selenium import webdriver
from selenium.webdriver import firefox
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from os import devnull

import re

from pyautogui import click
from pyautogui import press

from os import listdir
from openpyxl import load_workbook

from itertools import islice

from tkinter import Button, Tk, Label, Entry, BooleanVar, Checkbutton
# from pyautogui import click
from time import sleep

from datetime import datetime

from win32api import GetKeyState
from future.backports.http.client import GONE

from sys import argv
from sys import exit
from tkinter.constants import CURRENT
from _datetime import timedelta, date
from selenium.webdriver.firefox.options import Options
from HelperFunctions import done
import atexit
import os

# dayOfTheWeek = [-1]
# skip = [0]
# startAt = [""]

class AnyEc:
    """ Use with WebDriverWait to combine expected_conditions
        in an OR.
    """
    def __init__(self, *args):
        self.ecs = args
    def __call__(self, driver):
        for fn in self.ecs:
            try:
                if fn(driver): return True
            except:
                pass

class Container(object):
    containerNumber = ""
    PARS = ""
    driver = ""
    chassis = ""

def setupPortal(headlessInc):
    options = Options()
    options.set_headless(headless=headlessInc)
    fp = FirefoxProfile();
    fp.set_preference("webdriver.load.strategy", "unstable")
    
    downloadDirectory="J:\Linehaul\Linehaul Drivers Weekly Reports\ACI\\" + str(date.today())
    if not os.path.isdir(downloadDirectory):
        os.makedirs(downloadDirectory, exist_ok=True)

    #No download pop-up confirmation box
    fp.set_preference("browser.download.dir", downloadDirectory)
    fp.set_preference("browser.download.folderList", 2)
     
    fp.set_preference("browser.helperApps.neverAsk.saveToDisk", 
            "application/pdf");
    fp.set_preference("browser.download.manager.showWhenStarting", False)
    fp.set_preference("pdfjs.disabled", True)
    fp.set_preference("plugin.scan.plid.all", False)
    fp.set_preference("plugin.scan.Acrobat", "99.0")
    
    fp.set_preference("plugin.disable_full_page_plugin_for_types", "application/pdf")
    driver = webdriver.Firefox(firefox_profile=fp, log_path=devnull, firefox_options=options)
    driver.get("http://www.cbsa-asfc.gc.ca/prog/manif/portal-portail/menu-eng.html")
    driver.maximize_window()
    driver.implicitly_wait(30)
    
    elem = driver.find_element_by_css_selector('a[href*="https://apps-cbsa-asfc.fjgc-gccf.gc.ca/LCS/?l=eng&t=https://apps.cbsa-asfc.gc.ca/GCKey"]')
    elem.click()
    
    elem = driver.find_element_by_id("token1")
    f=open(r"C:\Automation\CBSA Login.txt", 'r')
    read = f.readline()
    m = re.search("username: *", read)
    read = read[m.end():]
    elem.send_keys(read)
    
    elem = driver.find_element_by_id("token2")
    read = f.readline()
    m = re.search("password: *", read)
    read = read[m.end():]
    elem.send_keys(read)
    f.close()
    
    elem = driver.find_element_by_css_selector('[title="Connect to the GCKey service"]')
    elem.click()
    
    
    elem = driver.find_element_by_id("continue")
    elem.click()
    
    
    elem = driver.find_element_by_name("_acceptEvent")
    elem.click()
    
    return driver

def loadInfo(folderPath):
    linehaul =""
    imports = ""
    
    containers = []
    
    for x in listdir(folderPath):
        if not x[0]=="~":
            if "Linehaul" in x:
                linehaul = x
#             if "Imports" in x and "aci" in x:
            if "Imports" in x:
                imports = x
            
    
    imports = load_workbook(folderPath+"\\"+imports)
    colorSheet = imports['COLOUR KEY']
    imports = imports['Imports Sheet']
    
    containerCol = ""
    driverCol = ""
    parsCol = ""
    
    for cell in next(imports.rows):
        if cell.value == "Container":
            containerCol = cell.col_idx - 1
        elif cell.value == "DRIVER":
            driverCol = cell.col_idx - 1
        elif cell.value == "PARS NUMBER":
            parsCol = cell.col_idx - 1
    
    colors = ["","","","",""]
#     print('aaa')
#     print(colorSheet.max_row)
    for xRow in colorSheet.rows:
        for xCell in xRow:
#             print("a")
#             print(xCell.value)
            if xCell.value=="MONDAY":
                colors[0]=xCell.offset(0,2).fill
                break
            elif xCell.value=="TUESDAY":
                colors[1]=xCell.offset(0,2).fill
                break
            elif xCell.value=="WEDNESDAY":
                colors[2]=xCell.offset(0,2).fill
                break
            elif xCell.value=="THURSDAY":
                colors[3]=xCell.offset(0,2).fill
                break
            elif xCell.value=="FRIDAY":
                colors[4]=xCell.offset(0,2).fill
                break
    

    dayOfTheWeek = [-1]
    skip = [0]
    startAt = [""]
    top = Tk()
    top.config(bg="lavender")
    L1 = Label(top, text="Please select which container to start at or\n how many containers to skip as well as\n which day of the week to run on \n ")
    L1.config(font=("Courier", 16), bg="lavender")
    L1.grid(row=0, column=0, columnspan=5)
    L2 = Label(top, text="Start at container #:")
    L2.config(font=("Courier", 10), bg="lavender")
    L2.grid(row=1, column=0, columnspan = 2)
    E1 = Entry(top, bd = 5, width = 39)
    E1.grid(row=1, column=2, columnspan = 2)
    L3 = Label(top, text="OR")
    L3.grid(row=2, column=1, columnspan=2)
    L3.config(font=("Courier", 20), bg="lavender")
    L4 = Label(top, text="# of containers to skip:")
    L4.grid(row = 3, column = 0, columnspan = 2)
    L4.config(font=("Courier", 10), bg="lavender")
    E2 = Entry(top, bd = 5, width = 39)
    E2.grid(row=3, column=2, columnspan=2)
    E2.insert(0, "0")
    top.lift()
    top.attributes('-topmost',True)
    top.after_idle(top.attributes,'-topmost',False)
      
    def callbackDay(day):
        startAt[0]=E1.get().strip()
        skip[0]=E2.get().strip()
        dayOfTheWeek[0] = day
        top.destroy()
      
      
    
    MyButton5 = Button(top, text="MONDAY", command=lambda: callbackDay(0))
    MyButton5.grid(row=4, column=0, padx = (20,0))
    MyButton5.config(font=("Courier", 16), bg="royal blue")
    MyButton6 = Button(top, text="TUESDAY", command=lambda: callbackDay(1))
    MyButton6.grid(row=4, column=1)
    MyButton6.config(font=("Courier", 16), bg="royal blue")
    MyButton7 = Button(top, text="WEDNESDAY",  command=lambda: callbackDay(2))
    MyButton7.grid(row=4, column=2)
    MyButton7.config(font=("Courier", 16), bg="royal blue")
    MyButton8 = Button(top, text="THURSDAY", command=lambda: callbackDay(3))
    MyButton8.grid(row=4, column=3)
    MyButton8.config(font=("Courier", 16), bg="royal blue")
    MyButton9 = Button(top, text="FRIDAY", command=lambda: callbackDay(4))
    MyButton9.grid(row=4, column=4, padx=(0,20))
    MyButton9.config(font=("Courier", 16), bg="royal blue")  
      
    checkHeadless = BooleanVar()
    checkHeadless.set(True)
    cb = Checkbutton(top, text="Run in background?", variable=checkHeadless, bg="brown1", font=("serif", 12))
    cb.grid(row=5, column=0, pady=(20), padx = 20, columnspan=5)  
      
    top.update()
    
    w = top.winfo_width() # width for the Tk root
    h = top.winfo_height() # height for the Tk root
       
    ws = top.winfo_screenwidth() # width of the screen
    hs = top.winfo_screenheight() # height of the screen
       
    # calculate x and y coordinates for the Tk root window
    x = (ws/2) - (w/2)
    y = (hs/2) - (h/2)
       
    # set the dimensions of the screen 
    # and where it is placed
    top.geometry('%dx%d+%d+%d' % (w, h, x, y))
   
    top.mainloop()
    
    
    
    todaysFill = colors[dayOfTheWeek[0]]
#    To do tomorrow's ACIs:
#     todaysFill = colors[datetime.today().weekday()+1]
#    To do Monday's ACIs on Friday:
#     todaysFill = colors[datetime.today().weekday()-4]
    #CHANGE THIS BACK
    startFound = False
    if startAt[0] != "":
        for row in imports.rows:
            if row[0].fill.fgColor == todaysFill.fgColor:
                if (not startFound) and row[containerCol].value==startAt[0]:
                    startFound = True
                if startFound:
                    container = Container()
                    container.containerNumber = str(row[containerCol].value).strip()
                    container.driver = str(row[driverCol].value).strip()
                    container.PARS = str(row[parsCol].value).strip()
                    containers.append(container)
                    
    elif skip[0]!= "0":
        i = 0
        for row in imports.rows:
            if row[0].fill.fgColor == todaysFill.fgColor:
                if i<int(skip[0]):
                    i+=1
                else:
                    container = Container()
                    container.containerNumber = str(row[containerCol].value).strip()
                    container.driver = str(row[driverCol].value).strip()
                    container.PARS = str(row[parsCol].value).strip()
                    containers.append(container)
    else:
        for row in imports.rows:
            if row[0].fill.fgColor == todaysFill.fgColor:
                container = Container()
                container.containerNumber = str(row[containerCol].value).strip()
                print(row[0].row)
                container.driver = str(row[driverCol].value).strip()
                container.PARS = str(row[parsCol].value).strip()
                containers.append(container)
    
    linehaul = load_workbook(folderPath+"\\"+linehaul)
    linehaul = linehaul['MASTER SHEET']
    
    chassisCol = ""
    driverLineCol = ""
    currentChassisCol = ''
#     next(linehaul.rows)
    for cell in next(islice(linehaul.rows, 0, 1)):
        if cell.value == "UNIT #":
            driverLineCol = cell.col_idx - 1
        elif cell.value == "ASSIGNED CHASSIS":
            chassisCol = cell.col_idx - 1
        elif cell.value == "CURRENTLY USING":
            currentChassisCol = cell.col_idx - 1
    
    containersToMatch = list(containers)
    
    for row in linehaul.rows:
        for container in containersToMatch:
            if str(row[driverLineCol].value) == container.driver or row[driverLineCol].value == container.driver:
                if str(row[currentChassisCol].value) != "" and str(row[currentChassisCol].value) != 'None':
                    container.chassis = str(row[currentChassisCol].value)
                    containersToMatch.remove(container)
                    break
                else:
                    container.chassis = str(row[chassisCol].value)
                    containersToMatch.remove(container)
                    break
    
    return containers, dayOfTheWeek[0], checkHeadless.get()
#     for container in containers:
#         print(container.containerNumber)
#         print(container.PARS)
#         print(container.driver)
#         print(container.chassis)
#     for i in range(imports.max_row):
        
            
def makeACI(driver, container, saved, failed, dayoftheweek):
    if (container.chassis=="" or
        container.driver=="" or
        container.containerNumber=="" or
        container.PARS==""):
        failed.append(container)
        return
    
    while not "Filter Submitted Documents list to view the following:" in driver.page_source:
        driver.find_element_by_id("tradeDocumentsTab").click()
    
    elem = driver.find_element_by_name("_create")
    elem.click()
    
    select = Select(driver.find_element_by_id("docTypeSelected"))
    select.select_by_visible_text("Highway Conveyance Document")
    
    elem = driver.find_element_by_id("submitButton")
    elem.click()
    
#     f=open(r"J:\Spencer\Current Year.txt", 'r')
#     year = f.readline()
#     f.close()
    year = datetime.now().year
    
#     print(datetime.now().weekday())
#     print(dayoftheweek)

    if dayoftheweek==0 and datetime.now().weekday() == 4:
        offsetDays = 3
    else:
        offsetDays = dayoftheweek - datetime.now().weekday()
    
    fiveDays = datetime.now() + timedelta(days=offsetDays)
        
    currentDate = str(fiveDays.year)
    
    if len(str(fiveDays.month))==1:
        currentDate+="0"
    currentDate+=str(fiveDays.month) 
    
    if len(str(fiveDays.day))==1:
        currentDate+="0" 
    currentDate+=str(fiveDays.day)
    
    driver.find_element_by_name("crnDocumentNumberForm.documentNumberWithoutClientCode").send_keys(container.containerNumber + '-' + str(year))
#     driver.find_element_by_id("datePicker").click()
#     click(683, 431)
    driver.find_element_by_name("conveyanceGeneralForm.estimatedDateOfArrival").send_keys(currentDate)
#     opened = False
#     driver.implicitly_wait(0)
#     while not opened:
#         try: 
#             driver.find_element_by_css_selector("td[class=' ui-datepicker-days-cell-over  ui-datepicker-today']").click()
#             opened = True
#         except:
# #             click(683, 431)
#             driver.find_element_by_name("conveyanceGeneralForm.estimatedDateOfArrival").click()
#     driver.implicitly_wait(600)
#     driver.find_element_by_css_selector("td[class=' ui-datepicker-days-cell-over  ui-datepicker-today']").click()
    
    driver.find_element_by_id("conveyanceGeneralForm.estimatedTimeOfArrivalHour").send_keys("17")
    driver.find_element_by_id("conveyanceGeneralForm.estimatedTimeOfArrivalMinute").send_keys("00")
    
    driver.find_element_by_id("portOfReport").send_keys("427")
    
    failed = True
    while failed:
        try:
            Select(driver.find_element_by_id("conveyanceGeneralForm.emptyLoadedStatusCode")).select_by_visible_text("Loaded")
            failed = False
        except:
            pass
    driver.find_element_by_id("conveyanceTractorTabTop").click()
    
    
    clicked = False
    while not clicked:
        try:
            driver.find_element_by_id("lookupName").clear()
            driver.find_element_by_id("lookupName").send_keys(str(container.driver))
            driver.find_element_by_id("lookupName").click()
            clicked = True
        except:pass
    

    driver.implicitly_wait(0)
    clicked = False
    while not clicked:
        try:
            elem = driver.find_element_by_css_selector("li[class='ui-menu-item']")
            clicked = True 
        except:
            pass
    
    while clicked:
        try:
            elem.click()
        except:
            clicked = False
    
    driver.implicitly_wait(30)
#     elem.send_keys(Keys.RETURN)
#     click(911, 468)

#     found = False
#     while not found:
#         try:
#             driver.find_element_by_css_selector("ul[class='ui-autocomplete ui-menu ui-widget ui-widget-content ui-corner-all']").click()
#             driver.find_element_by_css_selector("ul[class='ui-autocomplete ui-menu ui-widget ui-widget-content ui-corner-all']").click()
# #             sleep(10)
# #             press("enter")
#             found = True
#         except:
# #             click(911, 468)
#             driver.find_element_by_id("lookupName").clear()
#             driver.find_element_by_id("lookupName").send_keys(container.driver)
#             sleep(2)
#             
#     driver.implicitly_wait(600)
#     click(911, 468)
#     press("down")
#     press("enter")
#     PROBLEM MAYBE TRAVERSE THE TREE
#     driver.find_element_by_css_selector("a[innertext='" + str(container.driver) + "']").click()
#     sleep(1)
#     elem.click()
    while "Tractor Lookup Name" in driver.page_source:
        driver.find_element_by_id("conveyanceTrailerTabTop").click()
    driver.find_element_by_name("_enterTrailerInfo").click()
    driver.find_element_by_id("lookupName").send_keys(container.chassis)
#     WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "lookupName")))
    clicked = False
    while not clicked:
        try:
            driver.find_element_by_id("lookupName").click()
            clicked = True
        except:pass
#     sleep(1000)
#     driver.find_element_by_id("lookupName").click()
    
#     elem =driver.find_element_by_css_selector("ul[class='ui-autocomplete ui-menu ui-widget ui-widget-content ui-corner-all']") 
#     WebDriverWait(driver, 10).until(EC.element_to_be_clickable(elem))
#     elem.click()
    
#     click(891, 355)
#     driver.implicitly_wait(0)
#     found = False
#     while not found:
#         try:
#             driver.find_element_by_css_selector("ul[class='ui-autocomplete ui-menu ui-widget ui-widget-content ui-corner-all']").click()
#             driver.find_element_by_css_selector("ul[class='ui-autocomplete ui-menu ui-widget ui-widget-content ui-corner-all']").click()
# #             sleep(10)
# #             press("enter")
#             found = True
#         except:
# #             click(891, 355)
#             driver.find_element_by_id("lookupName").clear()
#             driver.find_element_by_id("lookupName").send_keys(container.chassis)
#             sleep(2)
# #     click(891, 355)
#     driver.implicitly_wait(600)
#     WebDriverWait(driver, 600).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "ul[class='ui-autocomplete ui-menu ui-widget ui-widget-content ui-corner-all']")))
#     press("down")
#     press("enter")
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "li[class='ui-menu-item']")))

    driver.implicitly_wait(0)
    clicked = False
    while not clicked:
        try:
            elem = driver.find_element_by_css_selector("li[class='ui-menu-item']")
            clicked = True 
        except:
            pass
    
    while clicked:
        try:
            elem.click()
        except:
            clicked = False
    
    driver.implicitly_wait(30)
    
    
    
    while "Trailer Lookup Name" in driver.page_source:
        driver.find_element_by_name("_enterContainerInfo").click()
    driver.find_element_by_name("id").send_keys(container.containerNumber)
    driver.find_element_by_name("ccns").send_keys(container.PARS)
    
    
    elem = driver.find_element_by_name("_save")
    elem.click()
    
    WebDriverWait(driver, 30).until(lambda driver: "Trailer Lookup Name" in driver.page_source)
    WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.NAME, "_save")))
#         sleep(5)

#         STILL A TIMING ISSUE
    
#     print("found")
#         WebDriverWait(driver, 10).until(EC.visibility_of_all_elements_located((By.CLASS_NAME, "buttonclass")))
#         print("found2")
#         if driver.current_url=="https://apps.cbsa-asfc.gc.ca/manif/services/eng/hb21006.html":
    while "Trailer Lookup Name" in driver.page_source:
        clicked = False
        while not clicked:
            try:
                elem = driver.find_element_by_name("_save")
                elem.click()
                clicked = True
            except:
                pass
#         else:s
#             sleep(0.1)
    elem = driver.find_element_by_name("_checkForErrors")
    elem.click()
    
    elem = driver.find_element_by_id("buttonPortalOk")
    elem.click()
    
    elem = driver.find_element_by_id("_submitToCBSA")
    elem.click()
    
    WebDriverWait(driver, 30).until(AnyEc(
        EC.presence_of_element_located((By.ID, "buttonPortalOk")),
        EC.presence_of_element_located((By.ID, "buttonPortalYes"))))
    
    driver.implicitly_wait(0)
    try:
        driver.find_element_by_id("buttonPortalOk").click()
        driver.implicitly_wait(30)
        failed.append(container)
        print("Failed: " + container.containerNumber)
        return
    except:
        driver.implicitly_wait(30)
        driver.find_element_by_id("buttonPortalYes").click()
        print("Created ACI for container: " + container.containerNumber)
        elem = driver.find_element_by_css_selector('a[href*="/manif/services/eng/pb10001.html?pb1=ls&ts="]')
        elem.click()
            
#         if not saved[0]:
#     #         while not GetKeyState(13)<0:
#     #             True
#             top = Tk()
#             L1 = Label()
#             L1 = Label(top, text="Check \"Do this automatically etc...\" and hit OK, then \n hit the \"CONTINUE\" button below")
#             L1.config(font=("Courier", 16))
#             L1.grid(row=0, column=0)
#             top.lift()
#             top.attributes('-topmost',True)
#             top.after_idle(top.attributes,'-topmost',False)
#               
#             def callbackOK():
#                 top.destroy()
#               
#             MyButton5 = Button(top, text="CONTINUE", command=callbackOK)
#             MyButton5.grid(row=1, column=0)
#               
#               
#             w = 500 # width for the Tk root
#             h = 150 # height for the Tk root
#                
#             # get screen width and height
#             ws = top.winfo_screenwidth() # width of the screen
#             hs = top.winfo_screenheight() # height of the screen
#                
#             # calculate x and y coordinates for the Tk root window
#             x = (ws/2) - (w/2)
#             y = (hs/2) - (h/2)
#                
#             # set the dimensions of the screen 
#             # and where it is placed
#             top.geometry('%dx%d+%d+%d' % (w, h, x, y))
#            
#             top.mainloop()
#         saved[0] = True
          
#         elem = driver.find_element_by_name("_ok")
#         elem.click()
#     click(535, 351)
#     

if __name__ == '__main__':
    argv = r"a J:\Linehaul\Linehaul Drivers Weekly Reports\2019\2019\Week 6".split()
#     argv = r"a C:\Users\ssleep\Documents\Week 29".split()
#     print("IF RUNNING IN BACKGROUND DO NOT EXIT THIS WINDOW")
#     print("HIT \"CONTROL-C\" TO END THE PROGRAM, AND THEN EXIT THE WINDOW") 
    
    folderPath = ''
    for i in range(len(argv)):
        if i!=0:
            folderPath+=argv[i]
            if i != len(argv) - 1:
                folderPath+=" "
    
#     driver = ""
    containers, dayoftheweek, headless = loadInfo(folderPath)
    driver = setupPortal(headless)
    saved = [False]
    failed = []
    i=0
    
    def exit_hander():
#         sleep(20)
        driver.quit()
     
    atexit.register(exit_hander)  
    
    for container in containers:
#         if i<21:
        try:
            makeACI(driver, container, saved, failed, dayoftheweek)
        except:
            print(container.containerNumber + "  FAILED")
            failed.append(container)
            driver.close()
            driver = setupPortal()
        i+=1
    
    failedFile = folderPath[:folderPath.rfind('\\')] + "\\" + "Failed Containers.txt"
    
    for container in failed:
        f=open(failedFile, "a+")
        f.write(container.containerNumber + "\n")
        f.close()
    
    done()
    driver.quit()
#     sleep(10000)
#     top = Tk()
#     L1 = Label()
#     L1 = Label(top, text="Please enter ERD:")
#     L1.grid(row=0, column=0)
#     E1 = Entry(top, bd = 5)
#     E1.grid(row=0, column=1)
    
#     def callbackNextACI(driver):
        
        
        
#     MyButton5 = Button(top, text="NextAci", height = 10, width=20, command=lambda: callbackNextACI(driver))
#     MyButton5.grid(row=0, column=0)
#     
#     top.lift()
#     top.attributes('-topmost',True)
#     top.after_idle(top.attributes,'-topmost',False)
#     
#     w = 200 # width for the Tk root
#     h = 100 # height for the Tk root
#     
#     # get screen width and height
#     ws = top.winfo_screenwidth() # width of the screen
#     hs = top.winfo_screenheight() # height of the screen
#     
#     # calculate x and y coordinates for the Tk root window
#     x = (ws/2) - (w/2)
#     y = (hs/2) - (h/2)
#     
#     # set the dimensions of the screen 
#     # and where it is placed
#     top.geometry('%dx%d+%d+%d' % (w, h, x, y))
# 
#     top.mainloop()

# pyinstaller "C:\Users\ssleep\workspace\New ACI\Automater\__init__.py" --distpath "J:\Spencer\CreateACIs" -y --noconsole