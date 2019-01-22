from selenium import webdriver
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile
from selenium.webdriver import Firefox
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select

from os import devnull
from os import system
# from win32com.client.gencache import EnsureDispatch #@UnresolvedImport
# from win32com.client import Dispatch #@UnresolvedImport
from openpyxl import load_workbook
from openpyxl.styles.colors import RGB
from openpyxl.styles import PatternFill
from copy import copy

from sys import path
from sys import exc_info
from time import sleep

from psutil import process_iter #@UnresolvedImport

from selenium.webdriver.common.action_chains import ActionChains

from tkinter import Button, Tk, Label, Entry

import re
from HelperFunctions import popUpOK
import sys

# from enum.IntFlag import  



class constants:
    CONTAINERCOL = ''
    STATUSCOL=""
    LOCCOL=""
    CUSTOMSCOL=""
    STEAMSHIPCOL=""
    ARRIVALDATECOL=""
    DEPARTUREDATECOL=""
    SIZECOL=""
    LFDCOL=""
    WEIGHTCOL=""
#     HAZCOL=""


class Container:
    def __init__(self, number = ""):
        self.number = number.upper()
        self.status = ""
        self.location = ""
        self.customs = ""
        self.steamship = ""
        self.arrivalDate =""
        self.departureDate =""
        self.size=""
        self.LFD=""
        self.weight=""


def setupTracer():
    fp = FirefoxProfile();
    fp.set_preference("webdriver.load.strategy", "unstable");
#     fp.set_preference("XRE_NO_WINDOWS_CRASH_DIALOG=1")
     
    driver = Firefox(firefox_profile=fp, log_path=devnull)
    driver.get("http://container-tracer.herokuapp.com/internaltools")
#     driver.set_window_position(1920, 0)
#     sleep(30)
    driver.maximize_window()
    
    driver.implicitly_wait(40)
    
    return driver
        
def readContainerInfo(driver, containers):
    driver.implicitly_wait(600)
#     driver.switch_to.frame("content3")
#     driver.switch_to.frame("main")
#     driver.switch_to_frame(driver.find_element_by_css_selector(""))
    
    containerText = driver.find_element_by_class_name("Textarea")
    for container in containers:
        containerText.send_keys(container.number + "\n")
        
    driver.find_element_by_id("submit").click()
    
    i = 0
#     driver.find_element_by_css_selector("tbody")
    
#     rows = driver.find_element_by_css_selector("table[class='data aligntop aligncenter']>tbody").find_elements_by_css_selector("tr")
    rows = driver.find_element_by_class_name("rt-tbody").find_elements_by_css_selector("div[role='row']")
    for row in rows:
        if i != 0:
            cells = row.find_elements_by_css_selector("td")
            container = ""
            matchString = cells[0].text
            while matchString[len(matchString)-1]==" ":
                matchString = matchString[:len(matchString)-2]
            m = re.compile(matchString[:4] + " ?0*" + matchString[5:], re.RegexFlag.IGNORECASE)
            for Xcontainer in containers:
#                 if Xcontainer.number[:4]+" "+Xcontainer.number[4:10] in cells[0].text:
                if m.search(Xcontainer.number):
                    container=Xcontainer
                    break
            j = 0
            if container != "":
                for cell in cells:
                    if j==1:
                        container.status = cell.text
                    elif j==2:
                        container.size = cell.text
                    elif j==3:
                        container.documents = cell.text
                    elif j==4:
                        container.customs = cell.text
                    elif j==5:
                        container.steamship = cell.text
                    elif j==6:
                        container.other = cell.text
                    elif j==7:
                        container.eta = cell.text
                    elif j==8:
                        container.storageDate = cell.text
                    elif j==9:
                        container.storageAmount = cell.text
                    j+=1
#                 else:
#                     print(container)
#                 print(cell.get_attribute("value"))
#                 print(cell.tag_name)
        i+=1
        
#     sleep(600)    
    driver.switch_to_default_content()
    driver.switch_to.frame("menuHeader")
#     driver.find_element_by_id("id41").click()
    driver.implicitly_wait(1)
    try:
        driver.find_element_by_id("id41").click()
    except:
        driver.implicitly_wait(600)
#         driver.find_element_by_id("tools").click()
        driver.find_element_by_id("tools").click()
        driver.find_element_by_id("tools").click()
        driver.find_element_by_id("tools").click()
#         sleep(5)
#         driver.find_element_by_class_name("tools selected").click()
        driver.switch_to_default_content()
        driver.switch_to_frame("content1")
#         sleep(500)
#         driver.find_element_by_css_selector(r'a[href^="top.frames[0].openTab(\'id41\');"]').click()
#         print(driver.find_element_by_css_selector('a[onclick*="top.frames[0].openTab(\'id41\');"]').text)
        driver.find_element_by_css_selector('a[onclick*="top.frames[0].openTab(\'id41\');"]').click()
        driver.switch_to_default_content()
        driver.switch_to.frame("menuHeader")
        driver.find_element_by_id("id41").click()


    driver.switch_to_default_content()
    driver.switch_to_frame(driver.find_element_by_css_selector("frame[name='content1']"))
    i = 2
    found = False
    driver.implicitly_wait(0)
    while not found:
        try:
            driver.switch_to_default_content()
            driver.switch_to_frame(driver.find_element_by_css_selector("frame[name='content" + str(i) + "']"))
            driver.find_element_by_css_selector("form[action='AppointmentQuery']")
            found = True
        except:
            if i<30:
                i+=1
            else:
                i=2
    driver.implicitly_wait(600)
    for k in range(int(len(containers)/20)+1):
        driver.find_element_by_css_selector("input[value='equipment']").click()
        containerText = driver.find_element_by_name("ids")
        m = k*20
        while m < (k+1)*20:
            if m<len(containers):
                noLeadingZeroes = containers[m].number
                while noLeadingZeroes[4]=="0":
                    noLeadingZeroes = noLeadingZeroes[0:4] + noLeadingZeroes[5:]
                containerText.send_keys(noLeadingZeroes + "\n")
            m+=1
            
        driver.find_element_by_id("btn_23").click()
        i=0
        table = driver.find_element_by_css_selector("table[class='TableStandardBG']")
        driver.implicitly_wait(0)
        try:
            rows = table.find_element_by_css_selector("table[id='listingTable']>tbody").find_elements_by_css_selector("tr")
            for row in rows:
                if i != 0:
                    cells = row.find_elements_by_css_selector("td")
                    matchString = cells[8].text
                    while matchString[len(matchString)-1]==" ":
                        matchString = matchString[:len(matchString)-2]
                    m = re.compile(matchString[:4] + " ?0*" + matchString[5:], re.RegexFlag.IGNORECASE)
                    if cells[4].text=="Active":
                        for Xcontainer in containers:
                            if m.search(Xcontainer.number):
                                container=Xcontainer
                                break
                        j = 0
                        if container != "":
            #                 for cell in cells:
            #                     print(cell.text)
            #                     if j==1:
                            container.rvtime = cells[2].text + ' ' + cells[3].text
            #                     elif j==2:
            #                     elif j==5:
                            container.rvnumber = cells[6].text
                i+=1
        except:
            pass
        driver.implicitly_wait(600)
        driver.find_element_by_css_selector("img[src='/ImxEbusWeb/images/english/Back.gif']").click() 
        
        
def putInfoinExcel(containers, localContainers):
    i = 0
    for row in localContainers.rows:
        if i != 0:
            if row[constants.CONTAINERCOL].value != None and row[constants.CONTAINERCOL].value != "":
                container = ""
                for Xcontainer in containers:
                    if Xcontainer.number==row[constants.CONTAINERCOL].value.upper():
                        container=Xcontainer
                        break
                    
                    
                backupRow = []
                j=0
                for _ in row:
                    if row[j].value == "":
                        backupRow.append(None)
                    else:
                        backupRow.append(row[j].value)
                    j+=1
                row[constants.STATUSCOL].value=container.status
                row[constants.SIZECOL].value=container.size
                row[constants.LOCCOL].value=container.location
                row[constants.CUSTOMSCOL].value=container.customs
                row[constants.STEAMSHIPCOL].value=container.steamship
                row[constants.ARRIVALDATECOL].value=container.arrivalDate
                row[constants.DEPARTUREDATECOL].value=container.departureDate
                row[constants.LFDCOL].value=container.LFD
                row[constants.WEIGHTCOL].value=container.weight
                
                
                hold = False
                for i in range(len(row)):
                    
        #             if backupRow[i].value != None and backupRow[i].value != row[i].value:
#                     if i>len(row)-4:
#                         print(backupRow[i])
#                         print(row[i].value)
#                         print(backupRow[i]==row[i].value)
                    if i>9:
                        if backupRow[i] != row[i].value and not (row[i].value=="" and backupRow[i]==None):
                            row[i].fill = PatternFill(start_color='CCFFFF',
                            end_color='CCFFFF',
                            fill_type='solid')
                        else:
                            row[i].fill = PatternFill(start_color='FFFFFF',
                            end_color='FFFFFF',
                            fill_type='solid')
                        
                if (row[constants.STEAMSHIPCOL].value and "Yes" in row[constants.STEAMSHIPCOL].value or row[constants.CUSTOMSCOL].value and "Yes" in row[constants.CUSTOMSCOL].value) and (row[constants.LOCCOL].value and row[constants.LOCCOL].value == "Seaport Intermodal"):
                    hold =True
                
                fillColor = "FFFFFF"
                
                if hold:
                    fillColor="FFC300"
                
                
                for j in range(3):
                    row[j].fill = PatternFill(start_color=fillColor,
                    end_color=fillColor,
                    fill_type='solid')
        
        i+=1
        dims = {}
        for row in localContainers.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        for col, value in dims.items():
            localContainers.column_dimensions[col].width = value+3
        
        
        
def setupExcel(filepPath, containers):
      
    localContainersWb = load_workbook(filepPath, read_only=False, keep_vba=True)
    localContainers = localContainersWb["Current"]
    
    for cell in range(1, localContainers.max_column):
#         print(cell)
        val = localContainers[1][cell].value
        if val=="Container": 
            constants.CONTAINERCOL=cell
        elif val=="Status": 
            constants.STATUSCOL=cell
        elif val=="Last known location": 
            constants.LOCCOL=cell
        elif val=="Customs release": 
            constants.CUSTOMSCOL=cell
        elif val=="Steamship release": 
            constants.STEAMSHIPCOL=cell
        elif val=="Arrival date": 
            constants.ARRIVALDATECOL=cell
        elif val=="Storage due": 
            constants.DEPARTUREDATECOL=cell
        elif val=="Departure date": 
            constants.RVNUMBERCOL=cell
        elif val=="Size": 
            constants.SIZECOL=cell
        elif val=="Last free day": 
            constants.LFDCOL=cell
        elif val=="Weight": 
            constants.WEIGHTCOL=cell
    
    for cell in range(2, localContainers.max_row+1):
        contNumber = localContainers[cell][int(constants.CONTAINERCOL)].value
#         print(contNumber)
        if contNumber != "" and contNumber != None:
            container = Container(contNumber)
            container.checkDigit=localContainers[cell][int(constants.CONTAINERCOL)+1].value
            containers.append(container)
    
    return localContainersWb, localContainers

# def copyRow(fromSheet, toSheet, indexCurrent, indexCompleted):
#     currentRow = fromSheet[indexCurrent]
#     completedRow = toSheet[indexCompleted]
#     for i in range(1, len(currentRow)):
#         new_cell=""
#         try:
#             new_cell = completedRow[i]
#         except:
#             new_cell = fromSheet.cell(row=indexCompleted, column=i)
#         cell=currentRow[i]
#         new_cell.value = currentRow[i].value
#         if cell.has_style:
#             new_cell.font = copy(cell.font)
#             new_cell.border = copy(cell.border)
#             new_cell.fill = copy(cell.fill)
#             new_cell.number_format = copy(cell.number_format)
#             new_cell.protection = copy(cell.protection)
#             new_cell.alignment = copy(cell.alignment)
    
# def moveCompleted(localContainersWb, filePath):
#     completed = localContainersWb["Completed"]
#     current = localContainersWb["Current"]
#     
# #     copyRow(current, completed, 1, 1)
#     destinationRow = completed.get_highest_row()
#     
#     for i in range(1, current.get_highest_row):
#         if "Load Out-Gate" in current.cell(row=i, column=7).value:
#             copyRow(current, completed, indexCurrent, indexCompleted) 
#     
#     
#     
#     dims = {}
#     for row in completed.rows:
#         for cell in row:
#             if cell.value:
#                 dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
#     for col, value in dims.items():
#         completed.column_dimensions[col].width = value+3
def checkContainerNumbers(containers):
    letterDictionary={"A":10,
                      "B":12,
                      "C":13,
                      "D":14,
                      "E":15,
                      "F":16,
                      "G":17,
                      "H":18,
                      "I":19,
                      "J":20,
                      "K":21,
                      "L":23,
                      "M":24,
                      "N":25,
                      "O":26,
                      "P":27,
                      "Q":28,
                      "R":29,
                      "S":30,
                      "T":31,
                      "U":32,
                      "V":34,
                      "W":35,
                      "X":36,
                      "Y":37,
                      "Z":38}
                      
    
    invalidContainerNumbers=""
    
    for container in containers:
        sumCheck=letterDictionary[container.number[0]]
        sumCheck+=letterDictionary[container.number[1]]*2
        sumCheck+=letterDictionary[container.number[2]]*4
        sumCheck+=letterDictionary[container.number[3]]*8
        sumCheck+=int(container.number[4])*16
        sumCheck+=int(container.number[5])*32
        sumCheck+=int(container.number[6])*64
        sumCheck+=int(container.number[7])*128
        sumCheck+=int(container.number[8])*256
        sumCheck+=int(container.number[9])*512
        
        check = sumCheck%11
        if check==10:
            check=0
#         print(container.number)
        if not check==int(container.checkDigit):
            invalidContainerNumbers+= container.number + "\n"
        
    if invalidContainerNumbers!="":
#         popUpOK("Invalid container numbers (by check digit):\n"+invalidContainerNumbers)
#         sys.exit()
        
        
        top = Tk()
        top.config(bg = "lavender")
        L1 = Label(top, text="Invalid container numbers (by check digit):\n"+invalidContainerNumbers, bg="lavender", font=("serif", 16))
        L1.grid(row=0, column=0, columnspan=2)
        
        def callbackEnd():
            sys.exit()
             
        
        def callbackContinue():
            top.destroy()
             
        
        MyButton4 = Button(top, text="Proceed anyway", width=14, command=lambda: callbackContinue(), bg="green", font=("serif", 16))
        MyButton4.grid(row=2, column=0, padx=10, pady=10)
        
        MyButton5 = Button(top, text="Stop", width=25, command=lambda: callbackEnd(), bg="red", font=("serif", 16))
        MyButton5.grid(row=2, column=1, padx=10, pady=10)
        
        top.update()

        w = top.winfo_width() # width for the Tk root
        h = top.winfo_height() # height for the Tk root
           
        ws = top.winfo_screenwidth() # width of the screen
        hs = top.winfo_screenheight() # height of the screen
        x = (ws/2) - (w/2)
        y = (hs/2) - (h/2)
        
        top.geometry('%dx%d+%d+%d' % (w, h, x, y))
        
        top.update()

        top.lift()
        top.attributes('-topmost',True)
        top.after_idle(top.attributes,'-topmost',False)
#         moveTo(MyButton5.winfo_width()/2 + MyButton5.winfo_rootx(), MyButton5.winfo_height()/2 + MyButton5.winfo_rooty())
        
        top.mainloop()
            
if __name__ == '__main__':
#     try:
    filePath = r"J:\LOCAL DEPARTMENT\Automation - DO NOT MOVE\Incoming Local Containers CY.xlsm"
#     filePath = r"C:\Incoming Local Containers.xlsm"
    containers = []
         
    localContainersWb, localContainers = setupExcel(filePath, containers)

    checkContainerNumbers(containers)
    
    driver = setupTracer()
    
    readContainerInfo(driver, containers)
    
    putInfoinExcel(containers, localContainers)
#     except:
#         print(exc_info())
#         sleep(60)
    saved = False
    while not saved:
        try:
            localContainersWb.save(filePath)
            saved = True
        except:
            top = Tk()
            L1 = Label()
            L1 = Label(top, text="Please close \"Incoming Local Containers\" \nspreadsheet then hit \"OK\"")
            L1.config(font=("Courier", 16))
            L1.grid(row=0, column=0)
            top.lift()
            top.attributes('-topmost',True)
            top.after_idle(top.attributes,'-topmost',False)
              
            def callbackOK():
                top.destroy()
              
            MyButton5 = Button(top, text="OK", width=8, command=callbackOK)
            MyButton5.grid(row=1, column=0)
            MyButton5.config(font=("Courier", 16))
              
              
            w = 530 # width for the Tk root
            h = 100 # height for the Tk root
               
            # get screen width and height
            ws = top.winfo_screenwidth() # width of the screen
            hs = top.winfo_screenheight() # height of the screen
               
            # calculate x and y coordinates for the Tk root window
            x = (ws/2) - (w/2)
            y = (hs/2) - (h/2)
               
            # set the dimensions of the screen 
            # and where it is placed
            top.geometry('%dx%d+%d+%d' % (w, h, x, y))
           
            top.mainloop()
#     driver.quit()
#     driver.find_element_by_tag_name('html').send_keys(Keys.CONTROL, Keys.SHIFT, "w")
#     ActionChains(driver).send_keys(Keys.CONTROL, Keys.SHIFT, "w").perform()

    driver.quit()
#     sleep(1)
#     while True:
#         for p in process_iter():
#             if "geckodriver.exe" in p.name():
#                 p.kill()
#             break
    system('start excel.exe "'+ filePath + '"')
    
# pyinstaller "C:\Users\ssleep\workspace\CNLister\Lister\__init__.py" --distpath "J:\Spencer\CNLister" --noconsole -y